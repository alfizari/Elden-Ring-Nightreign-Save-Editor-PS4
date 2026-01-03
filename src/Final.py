from main_file import decrypt_ds2_sl2, encrypt_modified_files
from main_file_import import decrypt_ds2_sl2_import
import json, shutil, os , struct
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from relic_checker import RelicChecker
from source_data_handler import SourceDataHandler
from typing import Optional


# Global variables
working_directory = os.path.dirname(os.path.abspath(__file__))
working_directory = Path(working_directory)
os.chdir(working_directory)

# Data storage
data_source = SourceDataHandler()
relic_checker: Optional[RelicChecker] = None
items_json = {}
effects_json = {}
data = None
userdata_path = None

# Config file path for remembering last opened file
CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "editor_config.json")

def load_config():
    """Load saved configuration"""
    try:
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, 'r') as f:
                return json.load(f)
    except:
        pass
    return {}

def save_config(config):
    """Save configuration to file"""
    try:
        with open(CONFIG_FILE, 'w') as f:
            json.dump(config, f, indent=2)
    except Exception as e:
        print(f"Warning: Could not save config: {e}")
imported_data=None
MODE = None
IMPORT_MODE=None
char_name_list = []
char_name_list_import = []
ga_relic = []
ga_items = []
ga_to_characters = {}  # Maps GA handle -> list of character names
current_murks = 0
current_sigs = 0
# AOB_search='00 00 00 00 ?? 00 00 00 ?? ?? 00 00 00 00 00 00 ??'
AOB_search='00 00 00 00 0A 00 00 00 ?? ?? 00 00 00 00 00 00 06'
from_aob_steam= 44
steam_id=None

# Character names for vessel assignment
CHARACTER_NAMES = ['Wylder', 'Guardian', 'Ironeye', 'Duchess', 'Raider',
                   'Revenant', 'Recluse', 'Executor', 'Scholar', 'Undertaker']

# Items type
ITEM_TYPE_EMPTY = 0x00000000
ITEM_TYPE_WEAPON = 0x80000000
ITEM_TYPE_ARMOR = 0x90000000
ITEM_TYPE_RELIC = 0xC0000000


def load_json_data():
    global items_json, effects_json
    try:
        # file_path = os.path.join(working_directory, "Resources/Json")

        # with open(os.path.join(file_path, 'items.json'), 'r', encoding='utf-8') as f:
        #     items_json = json.load(f)

        # with open(os.path.join(file_path, 'effects.json'), 'r', encoding='utf-8') as f:
        #     effects_json = json.load(f)
        items_json = data_source.get_relic_origin_structure()
        effects_json = data_source.get_effect_origin_structure()

        return True

    except FileNotFoundError as e:
        messagebox.showerror(
            "Error",
            f"JSON files not found: {str(e)}\nManual editing only available."
        )
        return False


def reload_language(language_code):
    global items_json, effects_json, data_source
    result = data_source.reload_text(language_code)
    items_json = data_source.get_relic_origin_structure()
    effects_json = data_source.get_effect_origin_structure()
    return result


class Item:
    BASE_SIZE = 8

    def __init__(self, gaitem_handle, item_id, effect_1, effect_2, effect_3,
                 durability, unk_1, sec_effect1, sec_effect2, sec_effect3,
                 unk_2, offset, extra=None, size=BASE_SIZE):
        self.gaitem_handle = gaitem_handle
        self.item_id = item_id
        self.effect_1 = effect_1
        self.effect_2 = effect_2
        self.effect_3 = effect_3
        self.durability = durability
        self.unk_1 = unk_1
        self.sec_effect1 = sec_effect1
        self.sec_effect2 = sec_effect2
        self.sec_effect3 = sec_effect3
        self.unk_2 = unk_2
        self.offset = offset
        self.size = size
        self.padding = extra or ()

    @classmethod
    def from_bytes(cls, data_type, offset=0):
        gaitem_handle, item_id = struct.unpack_from("<II", data_type, offset)
        type_bits = gaitem_handle & 0xF0000000
        cursor = offset + cls.BASE_SIZE
        size = cls.BASE_SIZE

        durability = unk_1 = unk_2 = 0
        effect_1 = effect_2 = effect_3 = 0
        sec_effect1 = sec_effect2 = sec_effect3 = 0
        padding = ()

        if gaitem_handle != 0:
            if type_bits == ITEM_TYPE_WEAPON:
                cursor += 80
                size = cursor - offset
            elif type_bits == ITEM_TYPE_ARMOR:
                cursor += 8
                size = cursor - offset
            elif type_bits == ITEM_TYPE_RELIC:
                durability, unk_1 = struct.unpack_from("<II", data_type, cursor)
                cursor += 8
                effect_1, effect_2, effect_3 = struct.unpack_from("<III", data_type, cursor)
                cursor += 12
                padding = struct.unpack_from("<7I", data_type, cursor)
                cursor += 0x1C
                sec_effect1, sec_effect2, sec_effect3 = struct.unpack_from("<III", data_type, cursor)
                cursor += 12
                unk_2 = struct.unpack_from("<I", data_type, cursor)[0]
                cursor += 12
                size = cursor - offset

        return cls(gaitem_handle, item_id, effect_1, effect_2, effect_3,
                   durability, unk_1, sec_effect1, sec_effect2, sec_effect3,
                   unk_2, offset, extra=padding, size=size)


def parse_items(data_type, start_offset, slot_count=5120):
    items = []
    offset = start_offset
    for _ in range(slot_count):
        item = Item.from_bytes(data_type, offset)
        items.append(item)
        offset += item.size
    return items, offset


def gaprint(data_type):
    global ga_relic, ga_items
    ga_items = []
    ga_relic = []
    start_offset = 0x14
    slot_count = 5120
    items, end_offset = parse_items(data_type, start_offset, slot_count)

    for item in items:
        type_bits = item.gaitem_handle & 0xF0000000
        ga_items.append((item.gaitem_handle, item.item_id, item.effect_1,
                        item.effect_2, item.effect_3, item.sec_effect1,
                        item.sec_effect2, item.sec_effect3, item.offset, item.size))
        
        if type_bits == ITEM_TYPE_RELIC:
            ga_relic.append((item.gaitem_handle, item.item_id, item.effect_1,
                           item.effect_2, item.effect_3, item.sec_effect1,
                           item.sec_effect2, item.sec_effect3, item.offset, item.size))
    return end_offset


def read_char_name(data):
    name_offset = gaprint(data) + 0x94
    max_chars = 16
    raw_name = data[name_offset:name_offset + max_chars * 2]
    name = raw_name.decode("utf-16-le", errors="ignore").rstrip("\x00")
    return name if name else None


def parse_vessel_assignments(file_data):
    """Parse vessel data to map GA handles to character names.

    Vessel structure:
    - Each character (0-9) has vessels at IDs 1000-1006, 2000-2006, etc.
    - Each vessel slot contains: ID (4 bytes) + 6 GA handles (24 bytes)
    """
    global ga_to_characters
    ga_to_characters = {}

    # Search for vessel structures in the save file
    for offset in range(0x10000, min(0x50000, len(file_data) - 28), 4):
        try:
            val = struct.unpack_from('<I', file_data, offset)[0]

            # Check for vessel IDs in range 1000-10006 (characters 0-9, vessels 0-6)
            if 1000 <= val <= 10006:
                char_id = (val - 1000) // 1000  # 0-based character index
                vessel_slot = val % 1000  # vessel 0-6

                # Only process valid vessel slots (0-6)
                if vessel_slot > 6:
                    continue

                # Read the 6 potential GA handles
                ga_handles = struct.unpack_from('<6I', file_data, offset + 4)

                for ga in ga_handles:
                    # Check if this is a relic GA handle (type bits = 0xC0000000)
                    if (ga & 0xF0000000) == ITEM_TYPE_RELIC and ga != 0:
                        char_name = CHARACTER_NAMES[char_id] if char_id < len(CHARACTER_NAMES) else f'Char_{char_id}'

                        if ga not in ga_to_characters:
                            ga_to_characters[ga] = set()
                        ga_to_characters[ga].add(char_name)
        except:
            continue

    # Convert sets to sorted lists for display
    for ga in ga_to_characters:
        ga_to_characters[ga] = sorted(list(ga_to_characters[ga]))

    return ga_to_characters


def read_murks_and_sigs(data):
    global current_murks, current_sigs
    offset = gaprint(data)
    name_offset = offset + 0x94
    murks_offset = name_offset + 52
    sigs_offset = name_offset - 64
    
    current_murks = struct.unpack_from('<I', data, murks_offset)[0]
    current_sigs = struct.unpack_from('<I', data, sigs_offset)[0]
    
    return current_murks, current_sigs


def write_murks_and_sigs(murks_value, sigs_value):
    global data
    offset = gaprint(data)
    name_offset = offset + 0x94
    murks_offset = name_offset + 52
    sigs_offset = name_offset - 64
    
    # Write murks
    murks_bytes = murks_value.to_bytes(4, 'little')
    data = data[:murks_offset] + murks_bytes + data[murks_offset+4:]
    
    # Write sigs
    sigs_bytes = sigs_value.to_bytes(4, 'little')
    data = data[:sigs_offset] + sigs_bytes + data[sigs_offset+4:]
    
    save_current_data()


def split_files(file_path, folder_name):
    file_name = os.path.basename(file_path)
    split_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), folder_name)
    #clean current dir
    if os.path.exists(split_dir):
        shutil.rmtree(split_dir)  # delete folder and everything inside
    os.makedirs(split_dir, exist_ok=True)

    if file_name.lower() == 'memory.dat':
        with open(file_path, "rb") as f:
            header = f.read(0x80)
            with open(os.path.join(split_dir, "header"), "wb") as out:
                out.write(header)
            
            chunk_size = 0x100000
            for i in range(10):
                data = f.read(chunk_size)
                if not data:
                    break
                with open(os.path.join(split_dir, f"userdata{i}"), "wb") as out:
                    data=bytearray(data)
                    data=(0x00100010).to_bytes(4, "little")+ data
                    out.write(data)
            
            regulation = f.read()
            if regulation:
                with open(os.path.join(split_dir, "regulation"), "wb") as out:
                    out.write(regulation)

    elif file_name == 'NR0000.sl2':
        decrypt_ds2_sl2(file_path)

def save_file():
    global data
    save_current_data()

    if MODE=='PC':

        output_sl2_file=filedialog.asksaveasfilename( initialfile="NR0000.sl2", title="Save PC SL2 save as")
        if not output_sl2_file:
            return
        
        encrypt_modified_files(output_sl2_file)

    if MODE == 'PS4':  ### HERE
        print('data length', len(data))
        
        # Validate data length before proceeding
        expected_length = 0x100004
        if len(data) != expected_length:
            messagebox.showerror('Error', 
                            f'Modified userdata size is invalid. '
                            f'Expected {hex(expected_length)}, got {hex(len(data))}. Cannot save.')
            return
        
        try:
            split_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'decrypted_output')
            
            # Validate split directory exists
            if not os.path.exists(split_dir):
                messagebox.showerror("Error", f"Directory not found: {split_dir}")
                return
            
            output_file = filedialog.asksaveasfilename(
                initialfile="memory.dat",
                title="Save PS4 save as",
                defaultextension=".dat",
                filetypes=[("DAT files", "*.dat"), ("All files", "*.*")]
            )
            
            if not output_file:
                return
            
            # Track total bytes written for validation
            total_bytes_written = 0
            
            with open(output_file, "wb") as out:
                # 1. HEADER
                header_path = os.path.join(split_dir, "header")
                if not os.path.exists(header_path):
                    messagebox.showerror("Error", f"Header file not found: {header_path}")
                    return
                
                with open(header_path, "rb") as f:
                    header_data = f.read()
                    if len(header_data) != 0x80:
                        messagebox.showerror("Error", 
                                        f"Invalid header size: {hex(len(header_data))}. "
                                        f"Expected {hex(0x80)} bytes.")
                        return
                    out.write(header_data)
                    total_bytes_written += len(header_data)
                
                print(f"Written header: {hex(total_bytes_written)} bytes")
                
                # 2. USERDATA 0–9
                check_padding = (0x00100010).to_bytes(4, "little")
                userdata_chunks_found = 0
                
                for i in range(10):
                    userdata_path = os.path.join(split_dir, f"userdata{i}")
                    
                    if not os.path.exists(userdata_path):
                        # Check if this is expected (some saves may have fewer chunks)
                        if i == 0:
                            messagebox.showerror("Error", 
                                            f"Required file not found: {userdata_path}")
                            return
                        else:
                            print(f"Warning: userdata{i} not found, stopping at {i} chunks")
                            break
                    
                    # Read original
                    with open(userdata_path, "rb") as f:
                        block = f.read()
                    
                    # Validate block has data
                    if len(block) < 4:
                        messagebox.showerror("Error", 
                                        f"userdata{i} is too small ({len(block)} bytes)")
                        return
                    
                    # PS4 USERDATA should start with 4 bytes padding
                    if block[:4] == check_padding:
                        # Strip the padding
                        block = block[4:]
                    else:
                        # Padding missing - this is suspicious but warn and continue
                        print(f"Warning: userdata{i} does not start with expected padding {check_padding.hex()}")
                        print(f"         Got: {block[:4].hex()}")
                        # Don't add padding, just use as-is
                    
                    # Validate chunk size (should be 0x100000 for full chunks)
                    expected_chunk_size = 0x100000
                    if len(block) != expected_chunk_size and i < 9:  # Last chunk might be smaller
                        print(f"Warning: userdata{i} has unexpected size {hex(len(block))}, "
                            f"expected {hex(expected_chunk_size)}")
                    
                    # Write block to output
                    out.write(block)
                    total_bytes_written += len(block)
                    userdata_chunks_found += 1
                
                print(f"Written {userdata_chunks_found} userdata chunks: {hex(total_bytes_written)} bytes total")
                
                # 3. REGULATION
                regulation_path = os.path.join(split_dir, "regulation")
                if os.path.exists(regulation_path):
                    with open(regulation_path, "rb") as f:
                        regulation_data = f.read()
                        if regulation_data:
                            out.write(regulation_data)
                            total_bytes_written += len(regulation_data)
                            print(f"Written regulation: {len(regulation_data)} bytes")
                        else:
                            print("Warning: regulation file is empty")
                else:
                    print("Warning: regulation file not found, skipping")
            
            # 4. SIZE VALIDATION
            final_size = os.path.getsize(output_file)
            expected_final_size = 0x12A00A0
            
            print(f"Final file size: {hex(final_size)} (expected: {hex(expected_final_size)})")
            
            if final_size != expected_final_size:
                messagebox.showerror('ERROR',
                                f'Invalid output file size!\n'
                                f'Expected: {hex(expected_final_size)} ({expected_final_size:,} bytes)\n'
                                f'Got: {hex(final_size)} ({final_size:,} bytes)\n'
                                f'Difference: {final_size - expected_final_size:+,} bytes\n\n'
                                f'File may be corrupt. Check the source files in {split_dir}')
                return
            
            messagebox.showinfo('Success', f'File saved successfully to:\n{output_file}')
            print(f"Successfully saved to: {output_file}")
            
        except PermissionError as e:
            messagebox.showerror("Permission Error", 
                            f"Cannot write to file. Check permissions.\n{str(e)}")
        except IOError as e:
            messagebox.showerror("I/O Error", 
                            f"Error reading/writing files.\n{str(e)}")
        except Exception as e:
            messagebox.showerror("Exception", 
                            f"Unexpected error occurred:\n{str(e)}\n\n"
                            f"Check console for details.")
            import traceback
            traceback.print_exc()

            

def name_to_path():
    global char_name_list, MODE
    char_name_list = []
    unpacked_folder = working_directory / 'decrypted_output'
    
    prefix = "userdata" if MODE == 'PS4' else "USERDATA_0"
    
    for i in range(10):
        file_path = os.path.join(unpacked_folder, f"{prefix}{i}")
        if not os.path.exists(file_path):
            continue
            
        try:
            with open(file_path, "rb") as f:
                file_data = f.read()
                name = read_char_name(file_data)
                if name:
                    char_name_list.append((name, file_path))
        except Exception as e:
            print(f"Error reading {file_path}: {e}")

def name_to_path_import():
    global char_name_list_import, IMPORT_MODE
    char_name_list_import = []
    unpacked_folder = working_directory / 'decrypted_output_import'
    
    prefix = "userdata" if IMPORT_MODE == 'PS4' else "USERDATA_0"
    
    for i in range(10):
        file_path = os.path.join(unpacked_folder, f"{prefix}{i}")
        if not os.path.exists(file_path):
            continue
            
        try:
            with open(file_path, "rb") as f:
                file_data = f.read()
                name = read_char_name(file_data)
                if name:
                    char_name_list_import.append((name, file_path))
        except Exception as e:
            print(f"Error reading {file_path}: {e}")


def delete_relic(ga_index, item_id):
    global data
    
    last_offset = gaprint(data)
    inventory_start = last_offset + 0x650
    inventory_end = inventory_start + 0xA7AB
    inventory_data = data[inventory_start:inventory_start + inventory_end]
    
    ga_bytes = ga_index.to_bytes(4, byteorder='little')
    replacement = bytes.fromhex('00000000FFFFFFFF')
    
    for ga, id, e1, e2, e3, e4, e5, e6, offset, size in ga_relic:
        real_id = id - 2147483648
        
        if ga_index == ga and real_id == item_id:
            inventory_offset = inventory_data.find(ga_bytes)
            match = inventory_offset + inventory_start
            
            data = data[:match] + b"\x00" * 14 + data[match+14:]

            data = data[:offset] + data[offset+80:]
            data = data[:offset] + replacement + data[offset:]

            data = data[:-0x1C] + b'\x00' * 72 + data[-0x1C:]
            
            save_current_data()
            return True
    return False


def modify_relic(ga_index, item_id, new_effects, new_item_id=None):
    global data

    for ga, id, e1, e2, e3, e4, e5, e6, offset, size in ga_relic:
        real_id = id - 2147483648

        if ga_index == ga and real_id == item_id:
            # Update item ID if provided
            if new_item_id is not None and new_item_id != real_id:
                # Convert to internal format
                new_id_internal = new_item_id + 2147483648
                item_id_offset = offset + 4  # Skip GA handle (4 bytes)
                item_id_bytes = new_id_internal.to_bytes(4, byteorder='little')
                data = data[:item_id_offset] + item_id_bytes + data[item_id_offset+4:]

            # Modify effects in the relic data structure
            effect_offset = offset + 16  # Skip handle, id, durability, unk_1

            # Write primary effects
            for i, eff in enumerate(new_effects[:3]):
                eff_bytes = eff.to_bytes(4, byteorder='little')
                pos = effect_offset + (i * 4)
                data = data[:pos] + eff_bytes + data[pos+4:]

            # Write secondary effects
            sec_effect_offset = effect_offset + 12 + 0x1C  # Skip padding
            for i, eff in enumerate(new_effects[3:6]):
                eff_bytes = eff.to_bytes(4, byteorder='little')
                pos = sec_effect_offset + (i * 4)
                data = data[:pos] + eff_bytes + data[pos+4:]

            save_current_data()
            return True
    return False


def modify_relic_by_ga(ga_index, new_effects, new_item_id, sort_effects=True):
    """Modify a relic by GA handle only (doesn't require matching item_id)"""
    global data

    for ga, id, e1, e2, e3, e4, e5, e6, offset, size in ga_relic:
        if ga_index == ga:
            # Sort effects if requested (fixes effect ordering issues)
            if sort_effects and relic_checker:
                new_effects = relic_checker.sort_effects(new_effects)

            # Update item ID
            new_id_internal = new_item_id + 2147483648
            item_id_offset = offset + 4  # Skip GA handle (4 bytes)
            item_id_bytes = new_id_internal.to_bytes(4, byteorder='little')
            data = data[:item_id_offset] + item_id_bytes + data[item_id_offset+4:]

            # Modify effects in the relic data structure
            effect_offset = offset + 16  # Skip handle, id, durability, unk_1

            # Write primary effects
            for i, eff in enumerate(new_effects[:3]):
                eff_bytes = eff.to_bytes(4, byteorder='little')
                pos = effect_offset + (i * 4)
                data = data[:pos] + eff_bytes + data[pos+4:]

            # Write secondary effects
            sec_effect_offset = effect_offset + 12 + 0x1C  # Skip padding
            for i, eff in enumerate(new_effects[3:6]):
                eff_bytes = eff.to_bytes(4, byteorder='little')
                pos = sec_effect_offset + (i * 4)
                data = data[:pos] + eff_bytes + data[pos+4:]

            save_current_data()
            return True
    return False


def check_illegal_relics():
    global relic_checker
    illegal_relics = relic_checker.illegal_gas
    # for ga, relic_id, e1, e2, e3, e4, e5, e6, offset, size in ga_relic:
    #     # Skip relic entirely if its ID is invalid
    #     if relic_id in (0, -1, 4294967295):
    #         continue

    #     effects = [e1, e2, e3, e4, e5, e6]
    #     used_ids = set()
    #     used_base_names = {}
    #     is_illegal = False

    #     for idx, eff in enumerate(effects, start=1):
    #         # Treat unknown effects as empty
    #         if eff in (0, -1, 4294967295):
    #             continue

    #         eff_key = str(eff)

    #         # Rule 1 — in illegal JSON
    #         if eff_key in ill_effects_json:
    #             is_illegal = True
    #             break

    #         # Lookup in main effects DB
    #         eff_name = effects_json.get(eff_key, {}).get("name", f"Unknown({eff})")

    #         # Rule 2 — duplicate ID
    #         if eff in used_ids:
    #             is_illegal = True
    #             break
    #         used_ids.add(eff)

    #         # Rule 3 — conflicting tiers
    #         base_name = eff_name.rsplit(" +", 1)[0] if " +" in eff_name else eff_name
    #         if base_name in used_base_names:
    #             is_illegal = True
    #             break

    #         used_base_names[base_name] = eff_name

    #     if is_illegal:
    #         illegal_relics.append(ga)

    return illegal_relics


def get_forbidden_relics():
    forbidden_relic_ids = RelicChecker.UNIQUENESS_IDS
    # forbidden_relic_ids = {
    #     1000, 1010, 1020, 1030, 1040, 1050, 1060, 1070, 1080, 1090,
    #     1100, 1110, 1120, 1130, 1140, 1150, 1160, 1170, 1180, 1190,
    #     1200, 1210, 1220, 1230, 1240, 1250, 1260, 1270, 11004, 10001,
    #     1400, 1410, 1420, 1430, 1440, 1450, 1460, 1470, 1480, 1490,
    #     1500, 1510, 1520
    # }
    return forbidden_relic_ids


def split_files_import(file_path, folder_name):
    global IMPORT_MODE
    file_name = os.path.basename(file_path)
    split_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), folder_name)
    # clean current dir
    if os.path.exists(split_dir):
        shutil.rmtree(split_dir)  # delete folder and everything inside
    os.makedirs(split_dir, exist_ok=True)

    if file_name.lower() == 'memory.dat':
        IMPORT_MODE = 'PS4'
        with open(file_path, "rb") as f:
            header = f.read(0x80)
            with open(os.path.join(split_dir, "header"), "wb") as out:
                out.write(header)
            
            chunk_size = 0x100000
            for i in range(10):
                data = f.read(chunk_size)
                if not data:
                    break
                with open(os.path.join(split_dir, f"userdata{i}"), "wb") as out:
                    data=bytearray(data)
                    data=(0x00100010).to_bytes(4, "little")+ data
                    out.write(data)
            
            regulation = f.read()
            if regulation:
                with open(os.path.join(split_dir, "regulation"), "wb") as out:
                    out.write(regulation)

    elif file_name == 'NR0000.sl2':
        IMPORT_MODE='PC'
        decrypt_ds2_sl2_import(file_path)

def import_save():
    global imported_data
    global char_name_list_import, data

    if data==None:
        messagebox.showerror('Error', 'Please select a character to replace first')
        return

    import_path = filedialog.askopenfilename()
    if not import_path:
        return

    # Split and generate list
    split_files_import(import_path, "decrypted_output_import")
    name_to_path_import()  # generates char_name_list_import = [(name, path), ...]

    # Show popup window with buttons
    show_import_popup()
    

def show_import_popup():
    popup = tk.Toplevel()
    popup.title("Select Character to Import")

    label = tk.Label(popup, text="Choose a character:", font=("Arial", 12, "bold"))
    label.pack(pady=10)

    # Create buttons for each character
    for name, path in char_name_list_import:
        btn = tk.Button(
            popup, 
            text=name, 
            width=30, 
            command=lambda p=path: load_imported_data_and_close(p, popup)
        )
        btn.pack(pady=3)

def load_imported_data_and_close(path, popup):
    load_imported_data(path)
    popup.destroy()

def load_imported_data(path):
    global imported_data, data

    with open (path, "rb") as f:
        imported_data=f.read()

    offsets = aob_search(imported_data, AOB_search)
    offset = offsets[0] + 44
    imported_data = imported_data[:offset] + bytes.fromhex(steam_id) + imported_data[offset + 8:] 


    if len(imported_data) <= len(data):
        data = imported_data + data[len(imported_data):]

    else:
        data = imported_data[:len(data)]

    for name, file in char_name_list_import:
        if path == file:
            char_name=name
    save_current_data()
    messagebox.showinfo("Success", f"Character '{char_name}' imported successfully. Save the file and open it again to see changes.")
    

    





    



def export_relics_to_excel(filepath="relics.xlsx"):
    if not ga_relic:
        return False, "No relics found in ga_relic."

    wb = Workbook()
    ws = wb.active
    ws.title = "Relics"

    # Column headers
    headers = [
        "Item ID",
        "Relic Name",
        "Relic Color",
        "Effect 1 (ID)", "Effect 1 (Name)",
        "Effect 2 (ID)", "Effect 2 (Name)",
        "Effect 3 (ID)", "Effect 3 (Name)",
        "Sec Effect 1 (ID)", "Sec Effect 1 (Name)",
        "Sec Effect 2 (ID)", "Sec Effect 2 (Name)",
        "Sec Effect 3 (ID)", "Sec Effect 3 (Name)",
    ]

    ws.append(headers)

    # Helper to fetch effect name
    def get_eff_name(eid):
        if eid in (0, -1, 4294967295):
            return "None"
        key = str(eid)
        return effects_json.get(key, {}).get("name", f"UnknownEffect({eid})")

    # Fill sheet
    for (_, item_id, e1, e2, e3, se1, se2, se3, offset, size) in ga_relic:
        # Skip invalid relics
        if item_id in (0, -1, 0xFFFFFFFF):
            continue

        real_id = item_id - 2147483648
        
        # Get relic name
        item_key = str(real_id)
        relic_name = items_json.get(item_key, {}).get("name", f"UnknownRelic({real_id})")
        relic_color = items_json.get(item_key, {}).get("color", "Unknown")

        row = [
            real_id,
            relic_name,
            relic_color,
            e1, get_eff_name(e1),
            e2, get_eff_name(e2),
            e3, get_eff_name(e3),
            se1, get_eff_name(se1),
            se2, get_eff_name(se2),
            se3, get_eff_name(se3),
        ]

        ws.append(row)

    # Auto-size columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max_len + 2

    try:
        wb.save(filepath)
        return True, f"Excel file saved: {filepath}"
    except Exception as e:
        return False, f"Failed to save Excel: {str(e)}"


def import_relics_from_excel(filepath):
    """
    Imports relics from an Excel file and modifies current relics to match.
    If imported list is longer than current relic list, extras are ignored.
    """
    global data, ga_relic
    
    if not ga_relic:
        return False, "No relics loaded in current save"
    
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Skip header row
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        modifications_made = 0
        
        # Iterate through current relics and imported data simultaneously
        for idx, (ga, item_id, e1, e2, e3, se1, se2, se3, offset, size) in enumerate(ga_relic):
            # If we've exhausted the imported list, stop
            if idx >= len(rows):
                break
            
            row = rows[idx]
            
            # Extract data from Excel row
            # Format: Item ID, Relic Name, Relic Color, E1(ID), E1(Name), E2(ID), E2(Name)...
            new_item_id = row[0] if row[0] is not None else 0
            new_e1 = row[3] if row[3] is not None else 0
            new_e2 = row[5] if row[5] is not None else 0
            new_e3 = row[7] if row[7] is not None else 0
            new_se1 = row[9] if row[9] is not None else 0
            new_se2 = row[11] if row[11] is not None else 0
            new_se3 = row[13] if row[13] is not None else 0
            
            # Convert item ID back to internal format
            new_item_id_internal = new_item_id + 2147483648
            
            # Modify the relic
            real_id = item_id - 2147483648
            
            # Update item ID if different
            if new_item_id != real_id:
                # Write new item ID
                item_id_offset = offset + 4  # Skip GA handle
                item_id_bytes = new_item_id_internal.to_bytes(4, byteorder='little')
                data = data[:item_id_offset] + item_id_bytes + data[item_id_offset+4:]
            
            # Update effects
            new_effects = [new_e1, new_e2, new_e3, new_se1, new_se2, new_se3]
            if modify_relic(ga, real_id, new_effects):
                modifications_made += 1
        
        save_current_data()
        return True, f"Successfully imported and modified {modifications_made} relic(s)"
        
    except Exception as e:
        return False, f"Failed to import Excel: {str(e)}"


def delete_all_illegal_relics():
    """Delete all relics with illegal effects"""
    global data
    
    illegal_gas = check_illegal_relics()
    
    if not illegal_gas:
        return 0, "No illegal relics found"
    
    deleted_count = 0
    failed_deletions = []
    
    for ga in illegal_gas:
        # Find the relic with this GA
        for ga_handle, id, e1, e2, e3, e4, e5, e6, offset, size in ga_relic:
            if ga_handle == ga:
                real_id = id - 2147483648
                if delete_relic(ga, real_id):
                    deleted_count += 1
                else:
                    failed_deletions.append(ga)
                break
    
    if failed_deletions:
        return deleted_count, f"Deleted {deleted_count} relics, but {len(failed_deletions)} failed"
    else:
        return deleted_count, f"Successfully deleted {deleted_count} illegal relics"


def save_current_data():
    global data, userdata_path
    if data and userdata_path:
        with open(userdata_path, 'wb') as f:

            f.write(data)

def aob_to_pattern(aob: str):

    parts = aob.split()
    pattern = bytearray()
    mask = bytearray()
    for p in parts:
        if p == "??":
            pattern.append(0x00)   # placeholder
            mask.append(0)         # 0 = wildcard (must NOT be 0x00)
        else:
            pattern.append(int(p, 16))
            mask.append(1)         # 1 = must match exactly
    return bytes(pattern), bytes(mask)


def aob_search(data: bytes, aob: str):
    pattern, mask = aob_to_pattern(aob)
    L = len(pattern)
    mv = memoryview(data)

    start = 0x58524  # skip below this offset
    end = len(data) - L + 1

    for i in range(start, end):

        # Check bytes
        for j in range(L):

            b = mv[i + j]

            # Non-wildcard: must match exactly
            if mask[j]:
                if b != pattern[j]:
                    break

            # Wildcard: 
            # 2025-12-28: Allow 0x00 to resolve Steam ID detection issues.
            # Narrowed down AOB_str (bytes 5 & 17 fixed) to prevent false positives.
            else:
                # if b == 0:  # Removed this restriction
                #     break
                continue

        else:
            # Inner loop did not break → MATCH FOUND
            return [i]

    return []


def find_steam_id(section_data):
    # # 假設你的 Steam ID 是 '76561198000000000' (17位數字)
    # # 先將它轉為 8 byte 的 little-endian 二進制格式 (這是 Steam ID 常見的儲存方式)
    # import struct
    # target_steam_id_hex = struct.pack('<Q', int(76561198013358313)).hex().upper() 
    # # 或者直接用你已知的 16進位 字串搜尋

    # # 搜尋 section_data 中你 ID 出現的所有位置
    # target_bytes = struct.pack('<Q', int(76561198013358313))
    # index = section_data.find(target_bytes)
    # print(f"你的 Steam ID 出現在偏移量: {hex(index)}")
    # if index != -1:
    #     search_start = index - 44
    #     actual_aob = section_data[search_start : search_start + 17].hex(' ').upper()
    #     print(f"預期 AOB 位置的實際數據為: {actual_aob}")
    #     print(f"原本定義的 AOB 模式為: 00 00 00 00 ?? 00 00 00 ?? ?? 00 00 00 00 00 00 ??")

    offsets = aob_search(section_data, AOB_search)
    offset = offsets[0] + 44
    steam_id = section_data[offset:offset+8]

    hex_str = steam_id.hex().upper()

    return hex_str

class SearchableCombobox(ttk.Frame):
    """A combobox with search functionality and manual entry"""
    def __init__(self, parent, values, **kwargs):
        super().__init__(parent)
        
        self.all_values = values
        self.var = tk.StringVar()
        
        # Entry widget for typing
        self.entry = ttk.Entry(self, textvariable=self.var, **kwargs)
        self.entry.pack(fill='x')
        
        # Listbox for suggestions
        self.listbox = tk.Listbox(self, height=6)
        self.listbox.pack(fill='both', expand=True)
        self.listbox.pack_forget()  # Hidden initially
        
        # Bind events
        self.entry.bind('<KeyRelease>', self.on_keyrelease)
        self.entry.bind('<FocusOut>', self.on_focusout)
        self.listbox.bind('<<ListboxSelect>>', self.on_select)
        self.listbox.bind('<FocusOut>', self.on_focusout)
        
        self.update_listbox(values)
    
    def on_keyrelease(self, event):
        # Filter values based on entry
        value = self.var.get().lower()
        
        if value == '':
            filtered = self.all_values
        else:
            filtered = [item for item in self.all_values if value in item.lower()]
        
        self.update_listbox(filtered)
        
        if filtered and event.keysym not in ('Up', 'Down', 'Return'):
            self.listbox.pack(fill='both', expand=True)
        elif not filtered:
            self.listbox.pack_forget()
    
    def update_listbox(self, values):
        self.listbox.delete(0, tk.END)
        for item in values:
            self.listbox.insert(tk.END, item)
    
    def on_select(self, event):
        if self.listbox.curselection():
            self.var.set(self.listbox.get(self.listbox.curselection()))
            self.listbox.pack_forget()
    
    def on_focusout(self, event):
        # Small delay to allow click on listbox
        self.after(100, lambda: self.listbox.pack_forget())
    
    def get(self):
        return self.var.get()
    
    def set(self, value):
        self.var.set(value)


class SaveEditorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Elden Ring NightReign Save Editor")
        self.root.geometry("1000x700")

        # Modify dialog reference
        self.modify_dialog = None

        # Clipboard for copy/paste relic effects
        self.clipboard_effects = None  # Will store (effects, item_id, item_name)

        # Track last selected character index for config saving
        self.last_char_index = None

        # Create notebook for tabs
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)

        # Create tabs
        self.file_tab = ttk.Frame(self.notebook)
        self.inventory_tab = ttk.Frame(self.notebook)

        self.notebook.add(self.file_tab, text="File Management")
        self.notebook.add(self.inventory_tab, text="Relics")

        self.setup_file_tab()
        self.setup_inventory_tab()

        # Try to load last opened file after UI is set up
        self.root.after(100, self.try_load_last_file)

    def try_load_last_file(self):
        """Try to load the last opened save file and character"""
        global MODE

        config = load_config()
        last_file = config.get('last_file')
        last_mode = config.get('last_mode')
        last_char_index = config.get('last_char_index', 0)

        if not last_file or not os.path.exists(last_file):
            return

        try:
            # Set mode
            MODE = last_mode

            # Split files
            split_files(last_file, 'decrypted_output')

            # Load JSON data
            if not load_json_data():
                return

            # Get character names
            name_to_path()

            # Display character buttons
            self.display_character_buttons()

            # Auto-select the last used character if valid
            if char_name_list and 0 <= last_char_index < len(char_name_list):
                name, path = char_name_list[last_char_index]
                self.on_character_click(last_char_index, path, name)

        except Exception as e:
            print(f"Could not auto-load last file: {e}")

    def setup_file_tab(self):
        # Main container
        container = ttk.Frame(self.file_tab)
        container.pack(fill='both', expand=True, padx=15, pady=15)
        
        # File loading section
        file_frame = ttk.LabelFrame(container, text="Save File Management", padding=15)
        file_frame.pack(fill='x', pady=(0, 15))
        
        info_label = ttk.Label(file_frame, text="Load your Elden Ring NightReign save file to begin editing", 
                              foreground='gray')
        info_label.pack(pady=(0, 10))
        
        ttk.Button(file_frame, text="📁 Open Save File", command=self.open_file, width=20).pack(pady=5)
        ttk.Button(file_frame, text="💾 Save Modified File", command=self.save_changes).pack(padx=5)
        ttk.Button(file_frame, text="💾 Import save (PC/PS4)", command=self.import_save_tk).pack(padx=5)
        
        # Stats section
        stats_frame = ttk.LabelFrame(container, text="Character Statistics", padding=15)
        stats_frame.pack(fill='x', pady=(0, 15))
        
        # Murks row
        murks_row = ttk.Frame(stats_frame)
        murks_row.pack(fill='x', pady=5)
        
        ttk.Label(murks_row, text="Murks:", font=('Arial', 11, 'bold'), width=15).pack(side='left')
        self.murks_display = ttk.Label(murks_row, text="N/A", font=('Arial', 11), foreground='blue')
        self.murks_display.pack(side='left', padx=10)
        ttk.Button(murks_row, text="Edit", command=self.modify_murks, width=10).pack(side='right', padx=5)
        
        # Sigs row
        sigs_row = ttk.Frame(stats_frame)
        sigs_row.pack(fill='x', pady=5)
        
        ttk.Label(sigs_row, text="Sigs:", font=('Arial', 11, 'bold'), width=15).pack(side='left')
        self.sigs_display = ttk.Label(sigs_row, text="N/A", font=('Arial', 11), foreground='blue')
        self.sigs_display.pack(side='left', padx=10)
        ttk.Button(sigs_row, text="Edit", command=self.modify_sigs, width=10).pack(side='right', padx=5)
        
        ttk.Button(stats_frame, text="🔄 Refresh Stats", command=self.refresh_stats).pack(pady=(10, 0))
        
        # Character selection section
        char_frame = ttk.LabelFrame(container, text="Select Character", padding=15)
        char_frame.pack(fill='both', expand=True)
        
        ttk.Label(char_frame, text="Choose a character to load:", foreground='gray').pack(anchor='w', pady=(0, 10))
        
        # Scrollable character list
        canvas = tk.Canvas(char_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(char_frame, orient="vertical", command=canvas.yview)
        self.char_button_frame = ttk.Frame(canvas)
        
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side='right', fill='y')
        canvas.pack(side='left', fill='both', expand=True)
        canvas.create_window((0, 0), window=self.char_button_frame, anchor='nw')
        
        self.char_button_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    

        
    def setup_inventory_tab(self):
        # Controls frame
        controls_frame = ttk.Frame(self.inventory_tab)
        controls_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Button(controls_frame, text="🔄 Refresh Inventory", command=self.refresh_inventory).pack(side='left', padx=5)
        ttk.Button(controls_frame, text="📤 Export to Excel", command=self.export_relics).pack(side='left', padx=5)
        ttk.Button(controls_frame, text="📥 Import from Excel", command=self.import_relics).pack(side='left', padx=5)
        ttk.Button(controls_frame, text="🗑️ Delete All Illegal", command=self.delete_all_illegal, 
                  style='Danger.TButton').pack(side='left', padx=5)
        ttk.Button(controls_frame, text="🗑️ Mass Delete Selected", command=self.mass_delete_relics,
                  style='Danger.TButton').pack(side='left', padx=5)
        ttk.Button(controls_frame, text="🔧 Mass Fix", command=self.mass_fix_incorrect_ids).pack(side='left', padx=5)
        
        # Info label
        self.illegal_count_label = ttk.Label(
            controls_frame,
            text="",
            foreground='red',
            font=('Arial', 9, 'bold')
        )
        self.illegal_count_label.pack(side='right', padx=10)

        legend_frame = ttk.Frame(controls_frame)
        legend_frame.pack(side='right', padx=10)

        ttk.Label(legend_frame, text="Blue = Red + Orange", foreground="blue").pack(side='left', padx=5)
        ttk.Label(legend_frame, text="Red = Illegal", foreground="red").pack(side='left', padx=5)
        ttk.Label(legend_frame, text="Purple = Missing Curse", foreground="#800080").pack(side='left', padx=5)
        ttk.Label(legend_frame, text="Orange = Unique Relic (don't edit)", foreground="#FF8C00").pack(side='left', padx=5)

        
        # Search frame
        search_frame = ttk.Frame(self.inventory_tab)
        search_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Label(search_frame, text="🔍 Search Relics:").pack(side='left', padx=5)
        self.search_var = tk.StringVar()
        self.search_var.trace('w', lambda *args: self.filter_relics())
        
        self.search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=30)
        self.search_entry.pack(side='left', padx=5)
        
        ttk.Button(search_frame, text="Clear", command=self.clear_search).pack(side='left', padx=5)

        # Character filter dropdown
        ttk.Label(search_frame, text="👤 Character:").pack(side='left', padx=(15, 5))
        self.char_filter_var = tk.StringVar(value="All")
        char_options = ["All"] + CHARACTER_NAMES
        self.char_filter_combo = ttk.Combobox(search_frame, textvariable=self.char_filter_var,
                                               values=char_options, state="readonly", width=12)
        self.char_filter_combo.pack(side='left', padx=5)
        self.char_filter_combo.bind("<<ComboboxSelected>>", lambda e: self.filter_relics())

        self.search_info_label = ttk.Label(search_frame, text="", foreground='gray')
        self.search_info_label.pack(side='left', padx=10)

        # ===========Add Language Combobox====================
        lang_frame = ttk.Frame(search_frame)
        lang_frame.pack(side='right', padx=5)

        ttk.Label(lang_frame, text="Language:").pack(side='left', padx=2)
        from source_data_handler import LANGUAGE_MAP
        lang_display_names = list(LANGUAGE_MAP.values())

        self.lang_combobox = ttk.Combobox(lang_frame,
                                          values=lang_display_names,
                                          state="readonly",
                                          width=15)
        self.lang_combobox.set(LANGUAGE_MAP.get("en_US"))
        self.lang_combobox.pack(side='left', padx=2)
        self.lang_combobox.bind("<<ComboboxSelected>>",
                                self.on_language_change)
        # ====================================================
        
        # Inventory display
        inv_frame = ttk.Frame(self.inventory_tab)
        inv_frame.pack(fill='both', expand=True, padx=10, pady=5)

        # Treeview for relics (added Equipped By and Deep columns)
        columns = ('Item Name', 'Deep', 'Item ID', 'Color', 'Equipped By', 'Effect 1', 'Effect 2', 'Effect 3',
                   'Sec Effect 1', 'Sec Effect 2', 'Sec Effect 3')

        self.tree = ttk.Treeview(inv_frame, columns=columns, show='tree headings', height=20)

        # Configure columns
        self.tree.heading('#0', text='#')
        self.tree.column('#0', width=40, minwidth=40, stretch=False)

        # Set column widths - more space for effect names
        col_widths = {
            'Item Name': 180,
            'Deep': 35,
            'Item ID': 80,
            'Color': 70,
            'Equipped By': 120,
            'Effect 1': 200,
            'Effect 2': 200,
            'Effect 3': 200,
            'Sec Effect 1': 200,
            'Sec Effect 2': 200,
            'Sec Effect 3': 200
        }

        for col in columns:
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_by_column(c))
            self.tree.column(col, width=col_widths.get(col, 150), minwidth=80)

        # Also allow sorting by # column
        self.tree.heading('#0', text='#', command=lambda: self.sort_by_column('#'))

        # Track sort state
        self.sort_column = None
        self.sort_reverse = False

        # Scrollbars
        vsb = ttk.Scrollbar(inv_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(inv_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Grid layout
        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        # Configure grid weights so scrollbars stay visible
        inv_frame.grid_rowconfigure(0, weight=1)
        inv_frame.grid_columnconfigure(0, weight=1)
        
        # Bind selection change
        self.tree.bind('<<TreeviewSelect>>', self.on_relic_select)

        # Bind double-click to open modify window
        self.tree.bind('<Double-1>', lambda e: self.modify_selected_relic())

        # Configure tree for extended selection (multiple items)
        self.tree.configure(selectmode='extended')

        # Context menu
        self.tree.bind("<Button-3>", self.show_context_menu)
        
        # Action buttons
        action_frame = ttk.Frame(self.inventory_tab)
        action_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Button(action_frame, text="Modify Selected", command=self.modify_selected_relic).pack(side='left', padx=5)
        ttk.Button(action_frame, text="Delete Selected", command=self.delete_selected_relic).pack(side='left', padx=5)
        
        # Selection controls
        selection_frame = ttk.Frame(action_frame)
        selection_frame.pack(side='left', padx=20)
        
        ttk.Button(selection_frame, text="Select All", command=self.select_all_relics).pack(side='left', padx=2)
        ttk.Button(selection_frame, text="Deselect All", command=self.deselect_all_relics).pack(side='left', padx=2)
        
        self.selection_count_label = ttk.Label(selection_frame, text="0 selected", foreground='blue', font=('Arial', 9, 'bold'))
        self.selection_count_label.pack(side='left', padx=10)
    
    
    def on_relic_select(self, event):
        """When a relic is selected and modify dialog is open, update the dialog"""
        if self.modify_dialog and self.modify_dialog.dialog.winfo_exists():
            selection = self.tree.selection()
            if selection:
                item = selection[0]
                tags = self.tree.item(item, 'tags')
                ga_handle = int(tags[0])
                item_id = int(tags[1])
                self.modify_dialog.load_relic(ga_handle, item_id)

    def on_language_change(self, event=None):
        selected_name = self.lang_combobox.get()
        from source_data_handler import LANGUAGE_MAP
        lang_code = next((code for code, name in LANGUAGE_MAP.items() if name == selected_name), "en_US")
        global data_source, items_json, effects_json
        if reload_language(lang_code):
            self.refresh_inventory()
            messagebox.showinfo("Success",
                                f"Language changed to: {selected_name}")
        else:
            messagebox.showerror("Error", "Can't change language.")

    def open_file(self):
        global MODE, data, userdata_path
        
        file_path = filedialog.askopenfilename(
            title="Select Save File",
        )
        
        if not file_path:
            return
        
        file_name = os.path.basename(file_path)
        
        # Determine mode
        if file_name.lower() == 'memory.dat':
            MODE = 'PS4'
            
        elif file_name == 'NR0000.sl2':
            MODE = 'PC'
        else:
            messagebox.showerror("Error", "If this is a PS4 save, make sure it is decrypted and change the file name to memory.dat")
            return
        
        # Split files
        split_files(file_path, 'decrypted_output')

        # Load JSON data
        if not load_json_data():
            return

        # Get character names
        name_to_path()

        # Save the opened file path to config
        config = load_config()
        config['last_file'] = file_path
        config['last_mode'] = MODE
        # Reset character index since we're opening a new file
        config['last_char_index'] = 0
        save_config(config)

        # Display character buttons
        self.display_character_buttons()
        
        
    def display_character_buttons(self):
        # Clear existing buttons
        for widget in self.char_button_frame.winfo_children():
            widget.destroy()
        
        # Create styles
        style = ttk.Style()
        style.configure("Char.TButton", font=('Arial', 10), padding=5)
        style.configure("Highlighted.TButton", font=('Arial', 10), padding=5, background="#AF4C4C", foreground="red")
        
        self.char_buttons = []
        
        columns = 4  # Number of buttons per row
        for idx, (name, path) in enumerate(char_name_list):
            row = idx // columns
            col = idx % columns
            
            # Button
            btn = ttk.Button(
                self.char_button_frame,
                text=f"{idx+1}. {name}",
                style="Char.TButton",
                command=lambda b_idx=idx, p=path, n=name: self.on_character_click(b_idx, p, n),
                width=20
            )
            btn.grid(row=row, column=col, padx=5, pady=5, sticky="nsew")
            self.char_buttons.append(btn)
        
        # Make columns expand evenly
        for col in range(columns):
            self.char_button_frame.grid_columnconfigure(col, weight=1)

    def on_character_click(self, idx, path, name):
        # Reset all buttons to normal style
        for b in self.char_buttons:
            b.configure(style="Char.TButton")

        # Highlight clicked button
        self.char_buttons[idx].configure(style="Highlighted.TButton")

        # Save selected character index to config
        self.last_char_index = idx
        config = load_config()
        config['last_char_index'] = idx
        save_config(config)

        # Load character
        self.load_character(path)

    def load_character(self, path):
        global data, userdata_path, steam_id, data_source, ga_relic, relic_checker, ga_to_characters
        userdata_path = path

        try:
            with open(path, "rb") as f:
                data = f.read()

            # Parse items
            gaprint(data)

            # Parse vessel assignments (maps GA handles to character names)
            parse_vessel_assignments(data)

            # Load Relic Checker
            relic_checker = RelicChecker(ga_relic=ga_relic,
                                         data_source=data_source)
            relic_checker.set_illegal_relics()
            # Read stats
            read_murks_and_sigs(data)

            steam_id = find_steam_id(data)

            # Refresh all tabs
            self.refresh_inventory()
            self.refresh_stats()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load character: {str(e)}")
    
    def refresh_stats(self):
        if data is None:
            return
        
        murks, sigs = read_murks_and_sigs(data)
        self.murks_display.config(text=str(murks))
        self.sigs_display.config(text=str(sigs))
    
    def modify_murks(self):
        if data is None:
            messagebox.showwarning("Warning", "No character loaded")
            return

        confrim = messagebox.askyesno("Confirm", "Modifying Murks would get you banned. Are you sure you want to proceed?")
        if not confrim:
            return
        
        new_value = simpledialog.askinteger("Modify Murks", 
                                           f"Current Murks: {current_murks}\n\nEnter new value (decimal):",
                                           initialvalue=current_murks)
        if new_value is not None:
            write_murks_and_sigs(new_value, current_sigs)
            self.refresh_stats()
            messagebox.showinfo("Success", "Murks updated successfully")
    
    def modify_sigs(self):
        if data is None:
            messagebox.showwarning("Warning", "No character loaded")
            return
        
        
        new_value = simpledialog.askinteger("Modify Sigs", 
                                           f"Current Sigs: {current_sigs}\n\nEnter new value (decimal):",
                                           initialvalue=current_sigs)
        if new_value is not None:
            write_murks_and_sigs(current_murks, new_value)
            self.refresh_stats()
            messagebox.showinfo("Success", "Sigs updated successfully")
    
    def refresh_inventory(self):
        global data, ga_relic, relic_checker

        if data is None:
            messagebox.showwarning("Warning", "No character loaded")
            return

        # Clear treeview
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Parse items - this updates ga_relic with current data
        gaprint(data)

        # Re-parse vessel assignments
        parse_vessel_assignments(data)

        # Update relic checker with new ga_relic and recalculate illegal relics
        if relic_checker:
            relic_checker.ga_relic = ga_relic
            relic_checker.set_illegal_relics()

        # Check for illegal relics
        illegal_gas = check_illegal_relics()
        illegal_count = len(illegal_gas)
        
        # Check for forbidden relics
        forbidden_relics = get_forbidden_relics()
        
        # Update illegal count label
        if illegal_count > 0:
            self.illegal_count_label.config(text=f"⚠️ {illegal_count} possible (test) Illegal Relic(s) Found")
        else:
            self.illegal_count_label.config(text="✓ All Relics Valid")
        
        # Store all relic data for filtering
        self.all_relics = []
        
        # Populate treeview
        for idx, (ga, id, e1, e2, e3, se1, se2, se3, offset, size) in enumerate(ga_relic):
            real_id = id - 2147483648
            
            # Get item name and color
            item_name = "Unknown"
            item_color = "Unknown"
            if str(real_id) in items_json:
                item_name = items_json[str(real_id)]["name"]
                item_color = items_json[str(real_id)].get("color", "Unknown")
            
            # Get effect names
            effects = [e1, e2, e3, se1, se2, se3]
            effect_names = []
            
            for eff in effects:
                if eff == 0:
                    effect_names.append("None")
                elif eff == 4294967295:
                    effect_names.append("Empty")
                elif str(eff) in effects_json:
                    effect_names.append(
                        "".join(effects_json[str(eff)]["name"].splitlines()))
                else:
                    effect_names.append(f"Unknown ({eff})")
            
            # Check if this relic is illegal or forbidden
            is_illegal = ga in illegal_gas
            is_forbidden = real_id in forbidden_relics
            is_curse_illegal = relic_checker and ga in relic_checker.curse_illegal_gas

            # Get character assignment (which characters have this relic equipped)
            equipped_by = ga_to_characters.get(ga, [])
            equipped_by_str = ", ".join(equipped_by) if equipped_by else "-"

            # Check if this is a deep relic (ID range 2000000-2019999)
            is_deep = 2000000 <= real_id <= 2019999

            # Determine tag
            tag_list = [ga, real_id]
            if is_forbidden and is_illegal:
                tag_list.append('both')
            elif is_forbidden:
                tag_list.append('forbidden')
            elif is_curse_illegal:
                tag_list.append('curse_illegal')
            elif is_illegal:
                tag_list.append('illegal')

            # Store relic data for filtering
            self.all_relics.append({
                'index': idx + 1,
                'ga': ga,
                'item_name': item_name,
                'real_id': real_id,
                'item_color': item_color,
                'is_deep': is_deep,
                'equipped_by': equipped_by,
                'equipped_by_str': equipped_by_str,
                'effect_names': effect_names,
                'tag_list': tuple(tag_list),
                'is_forbidden': is_forbidden,
                'is_illegal': is_illegal,
                'is_curse_illegal': is_curse_illegal,
                'both': is_forbidden and is_illegal
            })
        
        # Apply current filter (if any)
        self.filter_relics()
    
    def filter_relics(self):
        """Filter relics based on search term and character filter"""
        if not hasattr(self, 'all_relics'):
            return

        # Clear treeview
        for item in self.tree.get_children():
            self.tree.delete(item)

        search_term = self.search_var.get().lower()
        char_filter = self.char_filter_var.get() if hasattr(self, 'char_filter_var') else "All"

        # Filter relics
        filtered_relics = []
        for relic in self.all_relics:
            # Apply search filter
            passes_search = True
            if search_term != '':
                passes_search = (search_term in relic['item_name'].lower() or
                                 search_term in str(relic['real_id']) or
                                 search_term in relic.get('equipped_by_str', '').lower())

            # Apply character filter
            passes_char = True
            if char_filter != "All":
                equipped_by = relic.get('equipped_by', [])
                passes_char = char_filter in equipped_by

            if passes_search and passes_char:
                filtered_relics.append(relic)

        # Populate treeview with filtered results
        for relic in filtered_relics:
            deep_indicator = "✓" if relic.get('is_deep', False) else ""
            item_id = self.tree.insert('', 'end', text=str(relic['index']),
                           values=(relic['item_name'], deep_indicator, relic['real_id'], relic['item_color'],
                                  relic.get('equipped_by_str', '-'),
                                  relic['effect_names'][0], relic['effect_names'][1], relic['effect_names'][2],
                                  relic['effect_names'][3], relic['effect_names'][4], relic['effect_names'][5]),
                           tags=relic['tag_list'])

            # Color forbidden relics orange (priority over illegal)
            if relic['is_forbidden'] and relic['is_illegal']:
                self.tree.tag_configure('both', foreground='blue', font=('Arial', 9, 'bold'))
            elif relic['is_forbidden']:
                self.tree.tag_configure('forbidden', foreground='#FF8C00', font=('Arial', 9, 'bold'))
            # Color curse-illegal relics purple
            elif relic.get('is_curse_illegal', False):
                self.tree.tag_configure('curse_illegal', foreground='#9932CC', font=('Arial', 9, 'bold'))
            # Color illegal relics red
            elif relic['is_illegal']:
                self.tree.tag_configure('illegal', foreground='red', font=('Arial', 9, 'bold'))

        # Update search info
        filter_active = search_term or char_filter != "All"
        if filter_active:
            self.search_info_label.config(text=f"Showing {len(filtered_relics)} of {len(self.all_relics)} relics")
        else:
            self.search_info_label.config(text="")
    
    def clear_search(self):
        """Clear the search box and character filter"""
        self.search_var.set("")
        if hasattr(self, 'char_filter_var'):
            self.char_filter_var.set("All")
        self.filter_relics()
        self.search_entry.focus()

    def sort_by_column(self, col):
        """Sort treeview by clicked column"""
        if not hasattr(self, 'all_relics') or not self.all_relics:
            return

        # Toggle sort direction if same column clicked
        if self.sort_column == col:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = col
            self.sort_reverse = False

        # Define sort key based on column
        def get_sort_key(relic):
            if col == '#':
                return relic['index']
            elif col == 'Item Name':
                return relic['item_name'].lower()
            elif col == 'Deep':
                return relic.get('is_deep', False)
            elif col == 'Item ID':
                return relic['real_id']
            elif col == 'Color':
                return relic['item_color'].lower() if relic['item_color'] else ''
            elif col == 'Equipped By':
                return relic.get('equipped_by_str', '').lower()
            elif col == 'Effect 1':
                return relic['effect_names'][0].lower()
            elif col == 'Effect 2':
                return relic['effect_names'][1].lower()
            elif col == 'Effect 3':
                return relic['effect_names'][2].lower()
            elif col == 'Sec Effect 1':
                return relic['effect_names'][3].lower()
            elif col == 'Sec Effect 2':
                return relic['effect_names'][4].lower()
            elif col == 'Sec Effect 3':
                return relic['effect_names'][5].lower()
            return 0

        # Sort the data
        self.all_relics.sort(key=get_sort_key, reverse=self.sort_reverse)

        # Re-apply filter to refresh display
        self.filter_relics()

        # Update column header to show sort indicator
        columns = ('Item Name', 'Deep', 'Item ID', 'Color', 'Equipped By', 'Effect 1', 'Effect 2', 'Effect 3',
                   'Sec Effect 1', 'Sec Effect 2', 'Sec Effect 3')
        for c in columns:
            indicator = ''
            if c == col:
                indicator = ' ▼' if self.sort_reverse else ' ▲'
            self.tree.heading(c, text=c + indicator)

        # Update # column header
        if col == '#':
            self.tree.heading('#0', text='#' + (' ▼' if self.sort_reverse else ' ▲'))
        else:
            self.tree.heading('#0', text='#')

    def show_context_menu(self, event):
        # Select item under cursor
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)

            menu = tk.Menu(self.root, tearoff=0)
            menu.add_command(label="Modify", command=self.modify_selected_relic)
            menu.add_command(label="Delete", command=self.delete_selected_relic)
            menu.add_separator()
            menu.add_command(label="📋 Copy Effects", command=self.copy_relic_effects)

            # Only enable paste if we have something in clipboard
            paste_label = "📋 Paste Effects"
            if self.clipboard_effects:
                paste_label += f" (from {self.clipboard_effects[2]})"
            menu.add_command(label=paste_label, command=self.paste_relic_effects,
                           state='normal' if self.clipboard_effects else 'disabled')
            menu.post(event.x_root, event.y_root)

    def copy_relic_effects(self):
        """Copy effects from selected relic to clipboard"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "No relic selected")
            return

        item = selection[0]
        tags = self.tree.item(item, 'tags')
        ga_handle = int(tags[0])
        item_id = int(tags[1])

        # Find the relic data
        for ga, id, e1, e2, e3, se1, se2, se3, offset, size in ga_relic:
            real_id = id - 2147483648
            if ga == ga_handle and real_id == item_id:
                effects = [e1, e2, e3, se1, se2, se3]
                item_name = items_json.get(str(real_id), {}).get("name", f"Unknown ({real_id})")
                self.clipboard_effects = (effects, real_id, item_name)
                messagebox.showinfo("Copied", f"Copied effects from:\n{item_name}\n\nEffects: {len([e for e in effects if e != 0])} effect(s)")
                return

        messagebox.showerror("Error", "Could not find relic data")

    def paste_relic_effects(self):
        """Paste copied effects to selected relic"""
        if not self.clipboard_effects:
            messagebox.showwarning("Warning", "No effects copied. Right-click a relic and select 'Copy Effects' first.")
            return

        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "No relic selected")
            return

        item = selection[0]
        tags = self.tree.item(item, 'tags')
        ga_handle = int(tags[0])
        item_id = int(tags[1])

        # Get target relic info
        target_name = items_json.get(str(item_id), {}).get("name", f"Unknown ({item_id})")
        source_effects, source_id, source_name = self.clipboard_effects

        # Build effect names for display
        effect_names = []
        for eff in source_effects:
            if eff != 0:
                eff_name = effects_json.get(str(eff), {}).get("name", f"Unknown ({eff})")
                effect_names.append(eff_name)

        # Confirm paste
        msg = f"Paste effects from:\n  {source_name}\n\nTo:\n  {target_name}\n\n"
        msg += f"Effects to paste ({len(effect_names)}):\n"
        for name in effect_names[:6]:
            msg += f"  • {name}\n"
        msg += "\nProceed?"

        if not messagebox.askyesno("Confirm Paste", msg):
            return

        # Check if this would make the relic illegal
        if relic_checker:
            would_be_illegal = relic_checker.is_illegal(item_id, source_effects)
            if would_be_illegal:
                warn_msg = "⚠️ Warning: These effects may not be valid for this relic type.\n\n"
                warn_msg += "The relic may be flagged as illegal after pasting.\n\nContinue anyway?"
                if not messagebox.askyesno("Invalid Effects Warning", warn_msg, icon='warning'):
                    return

        # Apply the effects
        if modify_relic(ga_handle, item_id, source_effects):
            # Update illegal status
            if relic_checker:
                is_illegal = relic_checker.is_illegal(item_id, source_effects)
                is_curse_illegal = relic_checker.is_curse_illegal(item_id, source_effects)
                if is_illegal and ga_handle not in relic_checker.illegal_gas:
                    relic_checker.append_illegal(ga_handle, is_curse_illegal)
                elif not is_illegal and ga_handle in relic_checker.illegal_gas:
                    relic_checker.remove_illegal(ga_handle)

            messagebox.showinfo("Success", f"Effects pasted to {target_name}")
            self.refresh_inventory()
        else:
            messagebox.showerror("Error", "Failed to paste effects")

    def modify_selected_relic(self):
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "No relic selected")
            return
        
        item = selection[0]
        tags = self.tree.item(item, 'tags')
        ga_handle = int(tags[0])
        item_id = int(tags[1])
        
        # Check if this is a forbidden relic
        if 'forbidden' in tags:
            result = messagebox.askyesno(
                "⚠️ Warning - Do Not Edit Relic",
                f"This relic (ID: {item_id}) is flagged as 'Do Not Edit'.\n\n"
                "Modifying this relic may cause as ban\n"
                "Are you sure you want to proceed?",
                icon='warning'
            )
            if not result:
                return
        
        # If dialog doesn't exist or was closed, create new one
        if not self.modify_dialog or not self.modify_dialog.dialog.winfo_exists():
            self.modify_dialog = ModifyRelicDialog(self.root, ga_handle, item_id, self.refresh_inventory)
        else:
            # Update existing dialog with new relic
            self.modify_dialog.load_relic(ga_handle, item_id)
            self.modify_dialog.dialog.lift()
    
    def delete_selected_relic(self):
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "No relic selected")
            return
        
        # Check if multiple items selected
        if len(selection) > 1:
            result = messagebox.askyesno("Confirm Delete", 
                                         f"Are you sure you want to delete {len(selection)} relics?")
        else:
            item = selection[0]
            tags = self.tree.item(item, 'tags')
            item_id = int(tags[1])
            result = messagebox.askyesno("Confirm Delete", 
                                         f"Are you sure you want to delete this relic (ID: {item_id})?")
        
        if result:
            deleted_count = 0
            failed_count = 0
            
            for item in selection:
                tags = self.tree.item(item, 'tags')
                ga_handle = int(tags[0])
                item_id = int(tags[1])
                
                if delete_relic(ga_handle, item_id):
                    deleted_count += 1
                else:
                    failed_count += 1
            
            if deleted_count > 0:
                messagebox.showinfo("Success", f"Deleted {deleted_count} relic(s) successfully" + 
                                  (f"\n{failed_count} failed" if failed_count > 0 else ""))
                self.refresh_inventory()
            else:
                messagebox.showerror("Error", "Failed to delete relics")
    
    def select_all_relics(self):
        """Select all relics in the tree"""
        all_items = self.tree.get_children()
        self.tree.selection_set(all_items)
    
    def deselect_all_relics(self):
        """Deselect all relics"""
        self.tree.selection_remove(self.tree.selection())
    
    def invert_selection(self):
        """Invert the current selection"""
        all_items = self.tree.get_children()
        currently_selected = set(self.tree.selection())
        
        # Select items that aren't currently selected
        new_selection = [item for item in all_items if item not in currently_selected]
        
        self.tree.selection_remove(self.tree.selection())
        self.tree.selection_set(new_selection)
    
    def mass_delete_relics(self):
        """Delete all currently selected relics"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "No relics selected. Use the tree selection to choose relics to delete.")
            return
        
        # Check for forbidden relics in selection
        forbidden_count = 0
        for item in selection:
            tags = self.tree.item(item, 'tags')
            if 'forbidden' in tags:
                forbidden_count += 1
        
        # Confirmation message
        confirm_msg = f"Are you sure you want to delete {len(selection)} selected relic(s)?"
        if forbidden_count > 0:
            confirm_msg += f"\n\n⚠️ WARNING: {forbidden_count} of these are 'Do Not Edit' relics!"
            confirm_msg += "\n\nDeleting these may cause issues!"
        
        result = messagebox.askyesno("Confirm Mass Delete", confirm_msg, icon='warning' if forbidden_count > 0 else 'question')
        
        if not result:
            return
        
        # Delete all selected relics
        deleted_count = 0
        failed_count = 0
        
        for item in selection:
            tags = self.tree.item(item, 'tags')
            ga_handle = int(tags[0])
            item_id = int(tags[1])
            
            if delete_relic(ga_handle, item_id):
                deleted_count += 1
            else:
                failed_count += 1
        
        # Show result
        if deleted_count > 0:
            message = f"Successfully deleted {deleted_count} relic(s)"
            if failed_count > 0:
                message += f"\n{failed_count} failed to delete"
            messagebox.showinfo("Mass Delete Complete", message)
            self.refresh_inventory()
        else:
            messagebox.showerror("Error", "Failed to delete any relics")

    def mass_fix_incorrect_ids(self):
        """Find relics with incorrect IDs (effects don't match) and fix them by finding valid IDs"""
        global data, userdata_path

        if data is None:
            messagebox.showwarning("Warning", "No character loaded")
            return

        if relic_checker is None:
            messagebox.showwarning("Warning", "Relic checker not initialized")
            return

        # Find all illegal relics that could be fixed
        fixable_relics = []
        unfixable_relics = []  # Track relics that can't be fixed

        for ga, id, e1, e2, e3, se1, se2, se3, offset, size in ga_relic:
            real_id = id - 2147483648
            effects = [e1, e2, e3, se1, se2, se3]

            # Skip if not illegal
            if ga not in relic_checker.illegal_gas:
                continue

            # Skip unique relics
            if real_id in RelicChecker.UNIQUENESS_IDS:
                continue

            item_name = items_json.get(str(real_id), {}).get("name", f"Unknown ({real_id})")

            # Try to find a valid ID for these effects
            # This will find an ID with enough curse slots for effects that need curses
            valid_id = self._find_valid_relic_id_for_effects(real_id, effects)

            if valid_id is not None:
                # valid_id == real_id means effects are valid but might need reordering
                # valid_id != real_id means we need to change the item ID
                new_name = items_json.get(str(valid_id), {}).get("name", f"Unknown ({valid_id})")
                if valid_id == real_id:
                    # Same ID - just needs effect reordering
                    fixable_relics.append((ga, id, real_id, valid_id, item_name, f"{new_name} (reorder effects)", effects))
                else:
                    fixable_relics.append((ga, id, real_id, valid_id, item_name, new_name, effects))
            else:
                # Track why it can't be fixed
                reason = "No valid ID found with same color"
                unfixable_relics.append((real_id, item_name, reason))

        if not fixable_relics:
            msg = "No fixable relics found.\n\n"
            if unfixable_relics:
                msg += f"{len(unfixable_relics)} illegal relic(s) cannot be auto-fixed:\n"
                for real_id, name, reason in unfixable_relics[:5]:
                    msg += f"• {name} ({real_id}): {reason}\n"
                if len(unfixable_relics) > 5:
                    msg += f"... and {len(unfixable_relics) - 5} more\n"
                msg += "\nThese may need manual effect changes."
            messagebox.showinfo("Mass Fix", msg)
            return

        # Show confirmation with details
        details = f"Found {len(fixable_relics)} relic(s) that can be fixed:\n\n"
        for i, (ga, id, old_id, new_id, old_name, new_name, effects) in enumerate(fixable_relics[:10]):
            details += f"• {old_name} ({old_id}) → {new_name} ({new_id})\n"
        if len(fixable_relics) > 10:
            details += f"\n... and {len(fixable_relics) - 10} more"

        details += "\n\nProceed with fixing these relics?"

        result = messagebox.askyesno("Confirm Mass Fix", details)
        if not result:
            return

        # Apply fixes - use modify_relic_by_ga which matches by GA handle only
        # This avoids issues where item_id changes after gaprint refresh
        fixed_count = 0
        failed_count = 0

        for ga, id, old_id, new_id, old_name, new_name, effects in fixable_relics:
            # Reload ga_relic before each modification to ensure we have current offsets
            gaprint(data)

            # Use modify_relic_by_ga which only needs GA handle (doesn't care about current item_id)
            if modify_relic_by_ga(ga, effects, new_id):
                fixed_count += 1
            else:
                failed_count += 1

        # Reload data from file to ensure we have the final saved state
        if userdata_path:
            with open(userdata_path, 'rb') as f:
                data = f.read()

        # Show result
        message = f"Fixed {fixed_count} relic(s)"
        if failed_count > 0:
            message += f"\n{failed_count} failed to fix"

        messagebox.showinfo("Mass Fix Complete", message)
        self.refresh_inventory()

    def _find_valid_relic_id_for_effects(self, current_id, effects):
        """Find a valid relic ID that can have the given effects (must be same color)"""
        # Get the current relic's color
        if current_id not in data_source.relic_table.index:
            return None
        current_color = data_source.relic_table.loc[current_id, "relicColor"]

        # Count how many effects need curses
        curses_needed = sum(1 for e in effects[:3]
                          if e not in [0, -1, 4294967295] and data_source.effect_needs_curse(e))

        # Get the range group of current ID
        id_range = relic_checker.find_id_range(current_id)
        if not id_range:
            return None

        group_name, (range_start, range_end) = id_range

        # Skip illegal range
        if group_name == "illegal":
            return None

        # First check if current ID is actually valid (effects match AND has enough curse slots)
        # Use allow_empty_curses=True because we're checking if primary effects fit the pools
        if relic_checker._check_relic_effects_in_pool(current_id, effects, allow_empty_curses=True):
            # Also check it has enough curse slots for effects that need curses
            try:
                pools = data_source.get_relic_pools_seq(current_id)
                available_curse_slots = sum(1 for p in pools[3:] if p != -1)
                if available_curse_slots >= curses_needed:
                    return current_id  # Already valid, return same ID
            except KeyError:
                pass

        # Search within the same range for a valid ID with SAME color only
        for test_id in range(range_start, range_end + 1):
            if test_id not in data_source.relic_table.index:
                continue

            # Must be same color - never change color
            test_color = data_source.relic_table.loc[test_id, "relicColor"]
            if test_color != current_color:
                continue

            # Check if relic has enough curse slots for effects that need curses
            try:
                pools = data_source.get_relic_pools_seq(test_id)
                available_curse_slots = sum(1 for p in pools[3:] if p != -1)
                if available_curse_slots < curses_needed:
                    continue
            except KeyError:
                continue

            # Check if effects are valid for this ID (allow empty curses)
            if relic_checker._check_relic_effects_in_pool(test_id, effects, allow_empty_curses=True):
                return test_id

        return None

    def export_relics(self):
        """Export relics to Excel file"""
        if data is None:
            messagebox.showwarning("Warning", "No character loaded")
            return
        
        filepath = filedialog.asksaveasfilename(
            title="Save Excel File",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        
        if not filepath:
            return
        
        success, message = export_relics_to_excel(filepath)
        
        if success:
            messagebox.showinfo("Success", message)
        else:
            messagebox.showerror("Error", message)
    
    def import_relics(self):
        """Import relics from Excel file"""
        if data is None:
            messagebox.showwarning("Warning", "No character loaded")
            return
        
        filepath = filedialog.askopenfilename(
            title="Open Excel File",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        
        if not filepath:
            return
        
        # Confirm import
        result = messagebox.askyesno(
            "Confirm Import",
            "This will modify your current relics to match the imported file.\n\n"
            "• Relic IDs and effects will be replaced\n"
            "• Extra relics in the file will be ignored\n"
            "• Make sure you have a backup!\n\n"
            "Continue?"
        )
        
        if not result:
            return
        
        success, message = import_relics_from_excel(filepath)
        
        if success:
            messagebox.showinfo("Success", message)
            self.refresh_inventory()
        else:
            messagebox.showerror("Error", message)
    
    def delete_all_illegal(self):
        """Delete all relics with illegal effects"""
        if data is None:
            messagebox.showwarning("Warning", "No character loaded")
            return
        
        # Check for illegal relics first
        illegal_gas = check_illegal_relics()
        
        if not illegal_gas:
            messagebox.showinfo("Info", "No illegal relics found!")
            return
        
        # Confirm deletion
        result = messagebox.askyesno(
            "Confirm Deletion", 
            f"Found {len(illegal_gas)} illegal relic(s).\n\n"
            "This will permanently delete all relics with:\n"
            "• Effects in the illegal list\n"
            "• Duplicate effect IDs\n"
            "• Conflicting effect tiers\n\n"
            "Do you want to proceed?"
        )
        
        if not result:
            return
        
        # Delete all illegal relics
        count, message = delete_all_illegal_relics()
        
        if count > 0:
            messagebox.showinfo("Success", message)
            self.refresh_inventory()
        else:
            messagebox.showerror("Error", message)
    def import_save_tk(self):
        import_save()
        self.load_character(userdata_path)
    
    def save_changes(self):
        if data and userdata_path:
            save_file()
            messagebox.showinfo("Success", "Changes saved to file")
        else:
            messagebox.showwarning("Warning", "No character loaded")


class ModifyRelicDialog:
    def __init__(self, parent, ga_handle, item_id, callback):
        self.callback = callback
        
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Modify Relic")
        self.dialog.geometry("700x600")
        self.dialog.transient(parent)

        self.safe_mode_var = tk.BooleanVar(value=True)
        
        self.setup_ui()
        self.load_relic(ga_handle, item_id)
    
    def load_relic(self, ga_handle, item_id):
        """Load relic data into the dialog"""
        self.ga_handle = ga_handle
        self.item_id = item_id

        # Get current effects
        self.current_effects = self.get_current_effects()

        # Update UI
        self.update_effects_display()

        # Update debug info (after UI is set up)
        self.dialog.after(100, self.update_debug_info)
    
    def get_current_effects(self):
        for ga, id, e1, e2, e3, se1, se2, se3, offset, size in ga_relic:
            real_id = id - 2147483648
            if ga == self.ga_handle and real_id == self.item_id:
                return [e1, e2, e3, se1, se2, se3]
        return [0, 0, 0, 0, 0, 0]
    
    def update_effects_display(self):
        """Update the effect entry fields with current values"""
        # Update item ID display (use StringVar - this triggers _on_item_id_change)
        self.current_item_label.config(text=f"{self.item_id}")
        self.item_id_var.set(str(self.item_id))

        for i, entry in enumerate(self.effect_entries):
            current_eff = self.current_effects[i]
            if current_eff == 0:
                entry.delete(0, tk.END)
                entry.insert(0, "0")
            else:
                entry.delete(0, tk.END)
                entry.insert(0, str(current_eff))

                # Also update the name display
                if str(current_eff) in effects_json:
                    name = effects_json[str(current_eff)]["name"]
                    self.effect_name_labels[i].config(text=name)
                else:
                    self.effect_name_labels[i].config(text="Unknown Effect")

    def _update_color_display(self):
        """Update the current color label"""
        color_map = {0: ("Red", "#ff6666"), 1: ("Blue", "#6666ff"),
                     2: ("Yellow", "#cccc00"), 3: ("Green", "#66cc66")}
        try:
            current_relic_id = int(self.item_id_entry.get())
            color_idx = data_source.relic_table.loc[current_relic_id, "relicColor"]
            color_name, color_hex = color_map.get(color_idx, ("Unknown", "gray"))
            self.current_color_label.config(text=color_name, foreground=color_hex)
        except (KeyError, ValueError):
            self.current_color_label.config(text="Unknown", foreground="gray")

    def _update_relic_structure_display(self):
        """Update the relic structure label showing effect/curse slots"""
        try:
            current_relic_id = int(self.item_id_entry.get())
            pools = data_source.get_relic_pools_seq(current_relic_id)

            # Count effect slots and curse slots
            effect_slots = sum(1 for p in pools[:3] if p != -1)
            curse_slots = sum(1 for p in pools[3:] if p != -1)

            # Format the display text
            effect_text = f"{effect_slots} effect slot" + ("s" if effect_slots != 1 else "")
            curse_text = f"{curse_slots} curse slot" + ("s" if curse_slots != 1 else "")

            structure_text = f"({effect_text}, {curse_text})"
            self.relic_structure_label.config(text=structure_text)
        except (KeyError, ValueError):
            self.relic_structure_label.config(text="(Unknown structure)")

    def _on_item_id_change(self):
        """Called when item ID entry changes - updates color and structure display"""
        self._update_color_display()
        self._update_relic_structure_display()

    def update_debug_info(self):
        """Update debug info showing why relic is flagged"""
        if not hasattr(self, 'debug_text'):
            return

        self.debug_text.config(state='normal')
        self.debug_text.delete('1.0', tk.END)

        try:
            relic_id = int(self.item_id_entry.get())
            effects = []
            for entry in self.effect_entries:
                try:
                    effects.append(int(entry.get()))
                except ValueError:
                    effects.append(4294967295)

            lines = []
            lines.append(f"Relic ID: {relic_id}")
            lines.append(f"Effects: {effects[:3]}")
            lines.append(f"Curses:  {effects[3:]}")
            lines.append("")

            # Get relic pools
            try:
                pools = data_source.get_relic_pools_seq(relic_id)
                lines.append(f"Pools (eff1,eff2,eff3,curse1,curse2,curse3): {pools}")
            except KeyError:
                lines.append(f"Pools: Relic ID {relic_id} not found in table!")
                pools = None

            lines.append("")

            # Check validation status
            if relic_checker:
                is_illegal = relic_checker.is_illegal(relic_id, effects)
                is_curse_illegal = relic_checker.is_curse_illegal(relic_id, effects)

                lines.append(f"is_illegal(): {is_illegal}")
                lines.append(f"is_curse_illegal(): {is_curse_illegal}")
                lines.append("")

                # Show effect analysis
                lines.append("--- Effect Analysis ---")
                if pools:
                    effect_slot_count = sum(1 for p in pools[:3] if p != -1)
                    lines.append(f"Effect slots: {effect_slot_count} ({'single-effect' if effect_slot_count <= 1 else 'multi-effect'} relic)")

                    if effect_slot_count <= 1:
                        lines.append("Single-effect relics don't require curses")
                    else:
                        curse_required_count = 0
                        for i, eff in enumerate(effects[:3]):
                            if eff in [-1, 0, 4294967295]:
                                lines.append(f"Effect {i}: {eff} (empty)")
                            else:
                                needs_curse = data_source.effect_needs_curse(eff)
                                effect_pools = data_source.get_effect_pools(eff)
                                if needs_curse:
                                    curse_required_count += 1
                                lines.append(f"Effect {i}: {eff} -> needs_curse={needs_curse}, pools={effect_pools}")

                        curses_provided = sum(1 for c in effects[3:] if c not in [-1, 0, 4294967295])
                        lines.append("")
                        lines.append(f"Effects needing curses: {curse_required_count}, Curses provided: {curses_provided}")
                        if curse_required_count > curses_provided:
                            lines.append("⚠ NOT ENOUGH CURSES for effects that need them!")
                lines.append("")

                # Detailed check
                if pools:
                    effect_slot_count = sum(1 for p in pools[:3] if p != -1)
                    is_multi_effect = effect_slot_count > 1

                    lines.append("--- Sequence Check Details ---")
                    if not is_multi_effect:
                        lines.append("(Single-effect relic - curse checks skipped)")

                    possible_sequences = [[0, 1, 2], [0, 2, 1], [1, 0, 2],
                                          [1, 2, 0], [2, 0, 1], [2, 1, 0]]

                    for seq in possible_sequences:
                        cur_effects = [effects[i] for i in seq]
                        cur_curses = [effects[i + 3] for i in seq]

                        seq_valid = True
                        issues = []

                        for idx in range(3):
                            eff = cur_effects[idx]
                            curse = cur_curses[idx]
                            eff_pool = pools[idx]
                            curse_pool = pools[idx + 3]

                            # Check effect
                            if eff_pool == -1:
                                if eff != 4294967295:
                                    seq_valid = False
                                    issues.append(f"slot{idx}: pool=-1 but eff={eff}")
                            else:
                                pool_effs = data_source.get_pool_effects(eff_pool)
                                if eff not in pool_effs:
                                    seq_valid = False
                                    issues.append(f"slot{idx}: eff {eff} not in pool {eff_pool}")

                            # Check curse (only for multi-effect relics)
                            if is_multi_effect:
                                eff_needs_curse = data_source.effect_needs_curse(eff)

                                if eff_needs_curse:
                                    # Effect requires a curse
                                    if curse_pool == -1:
                                        issues.append(f"curse{idx}: effect needs curse but slot has no curse_pool!")
                                        seq_valid = False
                                    elif curse in [-1, 0, 4294967295]:
                                        issues.append(f"curse{idx}: effect REQUIRES curse, but empty!")
                                        seq_valid = False
                                    else:
                                        pool_curses = data_source.get_pool_effects(curse_pool)
                                        if curse not in pool_curses:
                                            issues.append(f"curse{idx}: {curse} not in pool {curse_pool}")
                                            seq_valid = False
                                elif curse_pool != -1:
                                    # Effect doesn't need curse but slot supports one (optional)
                                    if curse not in [-1, 0, 4294967295]:
                                        pool_curses = data_source.get_pool_effects(curse_pool)
                                        if curse not in pool_curses:
                                            issues.append(f"curse{idx}: {curse} not in pool {curse_pool}")
                                            seq_valid = False
                                else:
                                    # curse_pool == -1: no curse allowed
                                    if curse not in [-1, 0, 4294967295]:
                                        issues.append(f"curse{idx}: slot doesn't support curse but curse={curse}")
                                        seq_valid = False

                        status = "✓ VALID" if seq_valid else "✗ invalid"
                        lines.append(f"Seq {seq}: {status}")
                        if issues:
                            for issue in issues:
                                lines.append(f"  - {issue}")

            self.debug_text.insert('1.0', '\n'.join(lines))
        except Exception as e:
            self.debug_text.insert('1.0', f"Error generating debug info: {str(e)}")

        self.debug_text.config(state='disabled')

    def setup_ui(self):
        # Title showing current relic
        self.title_label = ttk.Label(self.dialog, text="", font=('Arial', 14, 'bold'))
        self.title_label.pack(pady=10)

        # Main container with scrollbar
        main_frame = ttk.Frame(self.dialog)
        main_frame.pack(fill='both', expand=True, padx=10, pady=6)

        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Enable mouse wheel scrolling only when mouse is over this dialog
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")

        def _bind_mousewheel(event):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)

        def _unbind_mousewheel(event):
            canvas.unbind_all("<MouseWheel>")

        # Bind on enter, unbind on leave
        self.dialog.bind("<Enter>", _bind_mousewheel)
        self.dialog.bind("<Leave>", _unbind_mousewheel)

        # Modifier Config Section
        modifier_frame = ttk.LabelFrame(scrollable_frame, text="Modifier Configuration", padding=10)
        modifier_frame.pack(fill='x', pady=5)
        self.safe_mode_cb = ttk.Checkbutton(modifier_frame,
                                            text="Safe Mode: Auto-filter legal effects",
                                            variable=self.safe_mode_var,
                                            onvalue=True, offvalue=False)
        self.safe_mode_cb.pack(anchor='w')
        
        # Item ID section (optional modification)
        item_frame = ttk.LabelFrame(scrollable_frame, text="Relic Item ID", padding=10)
        item_frame.pack(fill='x', pady=5)

        item_info_frame = ttk.Frame(item_frame)
        item_info_frame.pack(fill='x', anchor='w')

        ttk.Label(item_info_frame, text="Current Item ID:").pack(side='left')
        self.current_item_label = ttk.Label(item_info_frame, text="", font=('Arial', 10, 'bold'), foreground='blue')
        self.current_item_label.pack(side='left', padx=(5, 15))

        self.relic_structure_label = ttk.Label(item_info_frame, text="", font=('Arial', 9), foreground='#666666')
        self.relic_structure_label.pack(side='left')

        ttk.Label(item_frame, text="Enter new Item ID (decimal) or search:").pack(anchor='w', pady=(10, 0))

        item_entry_frame = ttk.Frame(item_frame)
        item_entry_frame.pack(fill='x', pady=5)

        self.item_id_var = tk.StringVar()
        self.item_id_var.trace('w', lambda *args: self._on_item_id_change())
        self.item_id_entry = ttk.Entry(item_entry_frame, width=15, textvariable=self.item_id_var)
        self.item_id_entry.pack(side='left', padx=5)

        ttk.Button(item_entry_frame, text="Search Items", command=self.search_items).pack(side='left', padx=5)
        ttk.Button(item_entry_frame, text="🔧 Find Valid ID", command=self.find_valid_relic_id).pack(side='left', padx=5)

        # Color change section
        color_frame = ttk.LabelFrame(scrollable_frame, text="Change Relic Color", padding=10)
        color_frame.pack(fill='x', pady=5)

        color_info_frame = ttk.Frame(color_frame)
        color_info_frame.pack(fill='x', pady=(0, 5))

        ttk.Label(color_info_frame, text="Current Color:").pack(side='left', padx=5)
        self.current_color_label = ttk.Label(color_info_frame, text="", font=('Arial', 10, 'bold'))
        self.current_color_label.pack(side='left', padx=5)

        color_btn_frame = ttk.Frame(color_frame)
        color_btn_frame.pack(fill='x', pady=5)

        ttk.Label(color_btn_frame, text="Change to:").pack(side='left', padx=5)

        # Color buttons with actual colors
        self.red_btn = tk.Button(color_btn_frame, text="🔴 Red", bg='#ffcccc', activebackground='#ff9999',
                                  command=lambda: self.change_relic_color(0))
        self.red_btn.pack(side='left', padx=3)

        self.blue_btn = tk.Button(color_btn_frame, text="🔵 Blue", bg='#ccccff', activebackground='#9999ff',
                                   command=lambda: self.change_relic_color(1))
        self.blue_btn.pack(side='left', padx=3)

        self.yellow_btn = tk.Button(color_btn_frame, text="🟡 Yellow", bg='#ffffcc', activebackground='#ffff99',
                                     command=lambda: self.change_relic_color(2))
        self.yellow_btn.pack(side='left', padx=3)

        self.green_btn = tk.Button(color_btn_frame, text="🟢 Green", bg='#ccffcc', activebackground='#99ff99',
                                    command=lambda: self.change_relic_color(3))
        self.green_btn.pack(side='left', padx=3)

        # Effect modification section
        effect_frame = ttk.LabelFrame(scrollable_frame, text="Modify Effects", padding=10)
        effect_frame.pack(fill='x', pady=5)
        
        self.effect_entries = []
        self.effect_name_labels = []
        effect_labels = ['Effect 1', 'Effect 2', 'Effect 3', 
                        'Secondary Effect 1', 'Secondary Effect 2', 'Secondary Effect 3']
        
        for i, label in enumerate(effect_labels):
            # Label
            ttk.Label(effect_frame, text=f"{label}:", font=('Arial', 10, 'bold')).grid(
                row=i*2, column=0, sticky='w', pady=(10, 2))
            
            # Entry frame
            entry_frame = ttk.Frame(effect_frame)
            entry_frame.grid(row=i*2+1, column=0, sticky='ew', pady=(0, 5))
            
            # Manual entry
            entry = ttk.Entry(entry_frame, width=15)
            entry.pack(side='left', padx=5)
            entry.bind('<KeyRelease>', lambda e, idx=i: self.on_effect_change(idx))
            self.effect_entries.append(entry)
            
            # Search button
            ttk.Button(entry_frame, text="Search Effects", 
                      command=lambda idx=i: self.search_effects(idx)).pack(side='left', padx=5)
            
            # Effect name display
            name_label = ttk.Label(entry_frame, text="", foreground='blue')
            name_label.pack(side='left', padx=5)
            self.effect_name_labels.append(name_label)
        
        effect_frame.grid_columnconfigure(0, weight=1)

        # Debug info section
        debug_frame = ttk.LabelFrame(scrollable_frame, text="⚙ Debug: Validation Info", padding=10)
        debug_frame.pack(fill='x', pady=5)

        self.debug_text = tk.Text(debug_frame, height=12, width=80, font=('Consolas', 9),
                                   state='disabled', bg='#f5f5f5')
        self.debug_text.pack(fill='x', padx=5, pady=5)

        ttk.Button(debug_frame, text="🔄 Refresh Debug Info",
                   command=self.update_debug_info).pack(anchor='w', padx=5)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Buttons at bottom
        button_frame = ttk.Frame(self.dialog)
        button_frame.pack(fill='x', padx=10, pady=10)

        ttk.Label(button_frame, text="Click different relics in inventory to switch",
                 foreground='gray').pack(side='left', padx=5)
        ttk.Button(button_frame, text="Apply Changes", command=self.apply_changes).pack(side='right', padx=5)
        ttk.Button(button_frame, text="Close", command=self.dialog.destroy).pack(side='right', padx=5)
        ttk.Button(button_frame, text="🔄 Auto Sort", command=self.auto_sort_effects).pack(side='right', padx=5)

    def auto_sort_effects(self):
        """Sort effects in the correct order (keeps curses with their corresponding effects)"""
        try:
            # Get current effects from entry fields
            current_effects = []
            for entry in self.effect_entries:
                try:
                    val = int(entry.get())
                    current_effects.append(val)
                except ValueError:
                    current_effects.append(4294967295)  # Empty

            if len(current_effects) != 6:
                messagebox.showerror("Error", "Invalid effects configuration")
                return

            # Use relic_checker to sort effects (keeps curses paired with their primary effects)
            if relic_checker:
                sorted_effects = relic_checker.sort_effects(current_effects)

                # Update entry fields with sorted values
                for i, entry in enumerate(self.effect_entries):
                    entry.delete(0, tk.END)
                    entry.insert(0, str(sorted_effects[i]))
                    self.on_effect_change(i)  # Update name labels

                messagebox.showinfo("Auto Sort", "Effects sorted successfully!\nCurses remain paired with their corresponding effects.")
            else:
                messagebox.showerror("Error", "Relic checker not initialized")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to sort effects: {str(e)}")

    def on_effect_change(self, index):
        """When effect ID is manually entered, update the name display"""
        try:
            effect_id = int(self.effect_entries[index].get())
            if effect_id == 0:
                self.effect_name_labels[index].config(text="None")
            elif str(effect_id) in effects_json:
                name = effects_json[str(effect_id)]["name"]
                self.effect_name_labels[index].config(text=name)
            else:
                self.effect_name_labels[index].config(text="Unknown Effect")
        except ValueError:
            self.effect_name_labels[index].config(text="Invalid ID")
    
    def search_items(self):
        """Open search dialog for items"""
        _items = {}
        if self.safe_mode_var.get():
            _cut_relic_id = int(self.item_id_entry.get())
            _, _safe_range = relic_checker.find_id_range(_cut_relic_id)
            _df = data_source.relic_table.copy()
            _df = _df[_df.index.isin(range(_safe_range[0], _safe_range[1] + 1))]
            _items = data_source.cvrt_filtered_relic_origin_structure(_df)
        else:
            _items = items_json
        SearchDialog(self.dialog, _items, "Select Relic", self.on_item_selected)

    def find_valid_relic_id(self):
        """Find a valid relic ID that matches the current effects configuration"""
        try:
            # Get current effects from the entry fields
            current_effects = []
            for entry in self.effect_entries:
                try:
                    val = int(entry.get())
                    current_effects.append(val)
                except ValueError:
                    current_effects.append(4294967295)  # Empty

            # Count how many effect slots are used (non-empty)
            effect_count = sum(1 for e in current_effects[:3] if e not in [0, -1, 4294967295])

            # Count how many curses are NEEDED based on the effects (not just present)
            # Effects that can only roll on 3-effect relics need curses
            curses_needed = sum(1 for e in current_effects[:3]
                               if e not in [0, -1, 4294967295] and data_source.effect_needs_curse(e))

            # Also count curses that are present (for validation)
            curses_present = sum(1 for e in current_effects[3:] if e not in [0, -1, 4294967295])

            # Get current relic's color
            current_relic_id = int(self.item_id_entry.get())
            try:
                current_color = data_source.relic_table.loc[current_relic_id, "relicColor"]
            except KeyError:
                current_color = None

            # Search for valid relics with matching structure
            relic_table = data_source.relic_table.copy()
            valid_candidates = []

            for relic_id, row in relic_table.iterrows():
                # Skip illegal range
                if relic_id in range(20000, 30036):
                    continue

                # Check if within valid range
                if relic_id not in range(100, 2013323):
                    continue

                # Get pool configuration for this relic
                pools = [
                    row["attachEffectTableId_1"],
                    row["attachEffectTableId_2"],
                    row["attachEffectTableId_3"],
                    row["attachEffectTableId_curse1"],
                    row["attachEffectTableId_curse2"],
                    row["attachEffectTableId_curse3"],
                ]

                # Count available slots (non -1 pools)
                available_effect_slots = sum(1 for p in pools[:3] if p != -1)
                available_curse_slots = sum(1 for p in pools[3:] if p != -1)

                # Check if this relic can accommodate our effects
                if available_effect_slots < effect_count:
                    continue

                # Must have enough curse slots for effects that NEED curses
                if available_curse_slots < curses_needed:
                    continue

                # Check color match (if we have a color preference)
                relic_color = row["relicColor"]
                if current_color is not None and relic_color != current_color:
                    continue

                # Check if all current effects are valid in this relic's pools
                effects_valid = self._check_effects_valid_for_relic(relic_id, current_effects, pools)
                if effects_valid:
                    valid_candidates.append(relic_id)

            if valid_candidates:
                # Pick the first valid candidate and update the entry
                best_match = valid_candidates[0]
                self.item_id_var.set(str(best_match))

                # Get the name for display
                relic_name = items_json.get(str(best_match), {}).get("name", "Unknown")
                relic_color = items_json.get(str(best_match), {}).get("color", "Unknown")

                messagebox.showinfo(
                    "Valid ID Found",
                    f"Found valid relic ID: {best_match}\n"
                    f"Name: {relic_name}\n"
                    f"Color: {relic_color}\n\n"
                    f"Effects: {effect_count}, Curses needed: {curses_needed}"
                )
            else:
                messagebox.showwarning(
                    "No Valid ID Found",
                    f"Could not find a valid relic ID that:\n"
                    f"- Has the same color\n"
                    f"- Supports {effect_count} effect slot(s)\n"
                    f"- Has at least {curses_needed} curse slot(s) for effects that need them\n"
                    f"- Has valid pools for all current effects\n\n"
                    f"Try removing some effects or changing the color requirement."
                )

        except Exception as e:
            messagebox.showerror("Error", f"Failed to find valid ID: {str(e)}")

    def _check_effects_valid_for_relic(self, relic_id, effects, pools):
        """Check if the given effects are valid for the relic's effect pools"""
        # For each non-empty effect, check if it exists in at least one valid pool
        for i, effect in enumerate(effects[:3]):
            if effect in [0, -1, 4294967295]:
                continue
            # Check if effect is in any of the effect pools
            found_in_pool = False
            for pool_id in pools[:3]:
                if pool_id == -1:
                    continue
                pool_effects = data_source.get_pool_effects(pool_id)
                if effect in pool_effects:
                    found_in_pool = True
                    break
            if not found_in_pool:
                return False

        # Check curse effects
        for i, effect in enumerate(effects[3:]):
            if effect in [0, -1, 4294967295]:
                continue
            # Check if curse effect is in any of the curse pools
            found_in_pool = False
            for pool_id in pools[3:]:
                if pool_id == -1:
                    continue
                pool_effects = data_source.get_pool_effects(pool_id)
                if effect in pool_effects:
                    found_in_pool = True
                    break
            if not found_in_pool:
                return False

        return True

    def change_relic_color(self, target_color):
        """Change the relic to a different color while keeping effects legal"""
        color_names = {0: "Red", 1: "Blue", 2: "Yellow", 3: "Green"}
        target_color_name = color_names.get(target_color, "Unknown")

        try:
            # Get current effects from the entry fields
            current_effects = []
            for entry in self.effect_entries:
                try:
                    val = int(entry.get())
                    current_effects.append(val)
                except ValueError:
                    current_effects.append(4294967295)  # Empty

            # Count how many effect slots are used (non-empty)
            effect_count = sum(1 for e in current_effects[:3] if e not in [0, -1, 4294967295])
            curse_count = sum(1 for e in current_effects[3:] if e not in [0, -1, 4294967295])

            # Get current relic's color to check if already the target
            current_relic_id = int(self.item_id_entry.get())
            try:
                current_color = data_source.relic_table.loc[current_relic_id, "relicColor"]
                if current_color == target_color:
                    messagebox.showinfo("Info", f"Relic is already {target_color_name}!")
                    return
            except KeyError:
                pass

            # Search for valid relics with the target color
            relic_table = data_source.relic_table.copy()
            valid_candidates = []

            for relic_id, row in relic_table.iterrows():
                # Skip illegal range
                if relic_id in range(20000, 30036):
                    continue

                # Check if within valid range
                if relic_id not in range(100, 2013323):
                    continue

                # Check color match
                if row["relicColor"] != target_color:
                    continue

                # Get pool configuration for this relic
                pools = [
                    row["attachEffectTableId_1"],
                    row["attachEffectTableId_2"],
                    row["attachEffectTableId_3"],
                    row["attachEffectTableId_curse1"],
                    row["attachEffectTableId_curse2"],
                    row["attachEffectTableId_curse3"],
                ]

                # Count available slots (non -1 pools)
                available_effect_slots = sum(1 for p in pools[:3] if p != -1)
                available_curse_slots = sum(1 for p in pools[3:] if p != -1)

                # Check if this relic can accommodate our effects
                if available_effect_slots < effect_count:
                    continue
                if available_curse_slots < curse_count:
                    continue

                # Check if all current effects are valid in this relic's pools
                effects_valid = self._check_effects_valid_for_relic(relic_id, current_effects, pools)
                if effects_valid:
                    valid_candidates.append(relic_id)

            if valid_candidates:
                # Pick the first valid candidate and update the entry
                best_match = valid_candidates[0]
                self.item_id_var.set(str(best_match))

                # Get the name for display
                relic_name = items_json.get(str(best_match), {}).get("name", "Unknown")

                messagebox.showinfo(
                    "Color Changed",
                    f"Changed to {target_color_name}!\n\n"
                    f"New Relic ID: {best_match}\n"
                    f"Name: {relic_name}\n\n"
                    f"Click 'Apply Changes' to save."
                )
            else:
                messagebox.showwarning(
                    "Cannot Change Color",
                    f"Could not find a valid {target_color_name} relic that:\n"
                    f"- Supports {effect_count} effect slot(s)\n"
                    f"- Supports {curse_count} curse slot(s)\n"
                    f"- Has valid pools for all current effects\n\n"
                    f"Try removing some effects first, or the effects may not exist in {target_color_name} relics."
                )

        except Exception as e:
            messagebox.showerror("Error", f"Failed to change color: {str(e)}")

    def search_effects(self, effect_index):
        """Open search dialog for effects"""
        _items = {}
        if self.safe_mode_var.get():
            _cut_relic_id = int(self.item_id_entry.get())
            _pool_id = data_source.get_relic_pools_seq(_cut_relic_id)[effect_index]
            _pool_effects = data_source.get_pool_effects(_pool_id)
            _effect_params_df = data_source.effect_params.copy()
            _effect_params_df = _effect_params_df[_effect_params_df.index.isin(_pool_effects)]
            match effect_index:
                case 1:
                    _effect_id_1 = int(self.effect_entries[0].get())
                    _conflic_id_1 = data_source.get_effect_conflict_id(_effect_id_1)
                    _effect_params_df = _effect_params_df[
                        (_effect_params_df["compatibilityId"] == -1) |
                        (_effect_params_df["compatibilityId"] != _conflic_id_1)
                    ]
                case 2:
                    _effect_id_1 = int(self.effect_entries[0].get())
                    _conflic_id_1 = data_source.get_effect_conflict_id(_effect_id_1)
                    _effect_id_2 = int(self.effect_entries[1].get())
                    _conflic_id_2 = data_source.get_effect_conflict_id(_effect_id_2)
                    _effect_params_df = _effect_params_df[
                        (_effect_params_df["compatibilityId"] == -1) |
                        ((_effect_params_df["compatibilityId"] != _conflic_id_1) &
                         (_effect_params_df["compatibilityId"] != _conflic_id_2))
                    ]
                case 4:
                    _effect_id_4 = int(self.effect_entries[3].get())
                    _conflic_id_4 = data_source.get_effect_conflict_id(_effect_id_4)
                    _effect_params_df = _effect_params_df[
                        (_effect_params_df["compatibilityId"] == -1) |
                        (_effect_params_df["compatibilityId"] != _conflic_id_4)
                    ]
                case 5:
                    _effect_id_4 = int(self.effect_entries[3].get())
                    _conflic_id_4 = data_source.get_effect_conflict_id(_effect_id_4)
                    _effect_id_5 = int(self.effect_entries[4].get())
                    _conflic_id_5 = data_source.get_effect_conflict_id(_effect_id_5)
                    _effect_params_df = _effect_params_df[
                        (_effect_params_df["compatibilityId"] == -1) |
                        ((_effect_params_df["compatibilityId"] != _conflic_id_4) &
                         (_effect_params_df["compatibilityId"] != _conflic_id_5))
                    ]
            _items = data_source.cvrt_filtered_effect_origin_structure(_effect_params_df)
        else:
            _items = effects_json
        SearchDialog(self.dialog, _items, f"Select Effect {effect_index + 1}", 
                    lambda item_id: self.on_effect_selected(effect_index, item_id))
    
    def on_item_selected(self, item_id):
        """Callback when item is selected from search"""
        self.item_id_var.set(str(item_id))
    
    def on_effect_selected(self, effect_index, effect_id):
        """Callback when effect is selected from search"""
        self.effect_entries[effect_index].delete(0, tk.END)
        self.effect_entries[effect_index].insert(0, str(effect_id))
        self.on_effect_change(effect_index)
    
    def apply_changes(self):
        # Extract effect IDs from entries
        global relic_checker
        new_effects = []
        
        for entry in self.effect_entries:
            try:
                value = int(entry.get())
                new_effects.append(value)
            except ValueError:
                new_effects.append(0)
        new_effects = relic_checker.sort_effects(new_effects)
        
        # Check if item ID was changed
        new_item_id = None
        try:
            entered_id = int(self.item_id_entry.get())
            if entered_id != self.item_id:
                new_item_id = entered_id
        except ValueError:
            pass  # Keep original ID if invalid entry
        
        is_illegal = relic_checker.is_illegal(entered_id, new_effects)
        is_curse_illegal = relic_checker.is_curse_illegal(entered_id, new_effects)
        if is_illegal and self.ga_handle not in relic_checker.illegal_gas:
            relic_checker.append_illegal(self.ga_handle, is_curse_illegal)
        elif not is_illegal and self.ga_handle in relic_checker.illegal_gas:
            relic_checker.remove_illegal(self.ga_handle)
        # Apply modifications
        if modify_relic(self.ga_handle, self.item_id, new_effects, new_item_id):
            messagebox.showinfo("Success", "Relic modified successfully")
            self.callback()
            # Update current item_id if it was changed
            if new_item_id is not None:
                self.item_id = new_item_id
            # Reload the current relic to show updated values
            self.load_relic(self.ga_handle, self.item_id)
        else:
            messagebox.showerror("Error", "Failed to modify relic")


class SearchDialog:
    """Search dialog for JSON items"""
    def __init__(self, parent, json_data, title, callback):
        self.json_data = json_data
        self.callback = callback
        
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("600x500")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        self.setup_ui()
    
    def setup_ui(self):
        # Search entry
        search_frame = ttk.Frame(self.dialog)
        search_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(search_frame, text="Search:").pack(side='left', padx=5)
        self.search_var = tk.StringVar()
        self.search_var.trace('w', lambda *args: self.filter_results())
        
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var)
        search_entry.pack(side='left', fill='x', expand=True, padx=5)
        search_entry.focus()
        
        # Results listbox
        results_frame = ttk.Frame(self.dialog)
        results_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        scrollbar = ttk.Scrollbar(results_frame)
        scrollbar.pack(side='right', fill='y')
        
        self.listbox = tk.Listbox(results_frame, yscrollcommand=scrollbar.set)
        self.listbox.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=self.listbox.yview)
        
        self.listbox.bind('<Double-Button-1>', self.on_select)
        
        # Populate initial results
        self.all_items = []
        for item_id, item_data in self.json_data.items():
            name = item_data.get('name', 'Unknown')
            self.all_items.append((item_id, name))
        
        self.all_items.sort(key=lambda x: int(x[0]) if x[0].isdigit() else 0)
        self.filter_results()
        
        # Buttons
        button_frame = ttk.Frame(self.dialog)
        button_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Button(button_frame, text="Select", command=self.on_select).pack(side='right', padx=5)
        ttk.Button(button_frame, text="Cancel", command=self.dialog.destroy).pack(side='right', padx=5)
    
    def filter_results(self):
        search_term = self.search_var.get().lower()
        
        self.listbox.delete(0, tk.END)
        
        for item_id, name in self.all_items:
            if search_term in name.lower() or search_term in item_id:
                self.listbox.insert(tk.END, f"{name} (ID: {item_id})")
    
    def on_select(self, event=None):
        selection = self.listbox.curselection()
        if not selection:
            return
        
        selected_text = self.listbox.get(selection[0])
        item_id = selected_text.split("ID: ")[1].rstrip(")")
        
        self.callback(int(item_id))
        self.dialog.destroy()


def main():
    root = tk.Tk()
    app = SaveEditorGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()

